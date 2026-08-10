"""File-system staging helpers for Crewdible transaction uploads."""

import zipfile
import xml.etree.ElementTree as ET
from io import BytesIO
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl
from sqlalchemy import text

from scripts.staging.manifest import (
    PROJECT_ROOT,
    calculate_sha256,
    ensure_file_manifest_tables,
    get_env_path,
    slugify,
    unique_path,
)


DEFAULT_CREWDIBLE_STAGING_ROOT = PROJECT_ROOT / "data" / "staging" / "crewdible"
CREWDIBLE_STAGING_ROOT = get_env_path(
    "CREWDIBLE_FILE_STAGING_ROOT",
    DEFAULT_CREWDIBLE_STAGING_ROOT,
)

CREWDIBLE_DATA_CATEGORY = "transaction"
XLSX_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

CREWDIBLE_REQUIRED_HEADERS = [
    "No",
    "Gudang",
    "Tanggal Transaksi",
    "No. Transaksi",
    "Status",
    "Nama Toko",
    "Nama Marketplace",
    "Nama Produk",
    "No. SKU",
    "Qty Produk",
    "Material Packaging",
    "Qty Material Packaging",
    "Total Biaya Transaksi",
]


@dataclass(frozen=True)
class CrewdibleUploadMetadata:
    period_year: int
    period_month: int | None = None
    data_category: str = CREWDIBLE_DATA_CATEGORY

    @property
    def period_label(self):
        if self.period_month:
            return f"{self.period_year}-{self.period_month:02d}"
        return str(self.period_year)


def inspect_crewdible_file(file_path):
    """Return a compact validity and row-count summary for a Crewdible workbook."""
    try:
        workbook_source = file_path
        if str(file_path).lower().endswith(".xls"):
            workbook_source = BytesIO(Path(file_path).read_bytes())
        wb = openpyxl.load_workbook(workbook_source, data_only=True, read_only=True)
        sheet_name = "Transaction" if "Transaction" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        rows = list(ws.values)
        wb.close()
    except Exception as exc:
        try:
            rows = _read_crewdible_xlsx_xml_rows(file_path)
            sheet_name = "Transaction"
        except Exception as fallback_exc:
            return {
                "valid": False,
                "sheet_name": "",
                "header_row": None,
                "physical_rows": 0,
                "transaction_rows": 0,
                "missing_headers": [],
                "error_message": f"Gagal membaca workbook: {exc}; fallback XML gagal: {fallback_exc}",
            }

    header_idx = None
    for idx, row in enumerate(rows[:10]):
        normalized = [str(c).strip() if c is not None else "" for c in row]
        if normalized and normalized[0] == "No" and "No. Transaksi" in normalized:
            header_idx = idx
            headers = normalized
            break
    else:
        headers = []

    if header_idx is None:
        return {
            "valid": False,
            "sheet_name": sheet_name,
            "header_row": None,
            "physical_rows": 0,
            "transaction_rows": 0,
            "missing_headers": CREWDIBLE_REQUIRED_HEADERS,
            "error_message": "Header Crewdible tidak ditemukan. Pastikan sheet Transaction memakai header di baris pertama.",
        }

    missing_headers = [h for h in CREWDIBLE_REQUIRED_HEADERS if h not in headers]
    no_index = headers.index("No") if "No" in headers else 0
    data_rows = rows[header_idx + 1 :]
    non_empty_rows = [row for row in data_rows if any(c is not None and str(c).strip() for c in row)]
    transaction_rows = [
        row for row in non_empty_rows
        if len(row) > no_index and row[no_index] is not None and str(row[no_index]).strip()
    ]

    return {
        "valid": not missing_headers,
        "sheet_name": sheet_name,
        "header_row": header_idx + 1,
        "physical_rows": len(non_empty_rows),
        "transaction_rows": len(transaction_rows),
        "missing_headers": missing_headers,
        "error_message": "" if not missing_headers else "Kolom wajib belum lengkap.",
    }


def _read_crewdible_xlsx_xml_rows(file_path):
    """Read xlsx-like files whose extension/content breaks openpyxl."""
    with zipfile.ZipFile(file_path, "r") as z:
        xml_raw = z.read("xl/worksheets/sheet1.xml").decode("utf-8")

    xml_raw = xml_raw.replace("<v>-</v>", "<v>0</v>")
    root = ET.fromstring(xml_raw)

    rows = []
    for row_el in root.iter(f"{{{XLSX_NS}}}row"):
        row_values = []
        for cell in row_el.iter(f"{{{XLSX_NS}}}c"):
            cell_type = cell.get("t", "n")
            if cell_type == "inlineStr":
                is_el = cell.find(f"{{{XLSX_NS}}}is")
                text_el = is_el.find(f"{{{XLSX_NS}}}t") if is_el is not None else None
                value = text_el.text if text_el is not None else None
            else:
                value_el = cell.find(f"{{{XLSX_NS}}}v")
                value = value_el.text if value_el is not None else None
            row_values.append(value)
        rows.append(tuple(row_values))
    return rows


def build_crewdible_staging_path(original_filename, metadata: CrewdibleUploadMetadata, uploaded_at=None):
    uploaded_at = uploaded_at or datetime.now()
    file_ext = Path(original_filename).suffix.lower() or ".xlsx"
    original_stem = slugify(Path(original_filename).stem)
    month_folder = f"month={metadata.period_month:02d}" if metadata.period_month else "month=all"

    folder = (
        CREWDIBLE_STAGING_ROOT
        / metadata.data_category
        / f"year={metadata.period_year}"
        / month_folder
        / f"uploaded_date={uploaded_at:%Y-%m-%d}"
    )
    folder.mkdir(parents=True, exist_ok=True)

    if metadata.period_month:
        period_part = f"{metadata.period_year}_{metadata.period_month:02d}"
    else:
        period_part = str(metadata.period_year)
    staged_name = f"crewdible_{metadata.data_category}_{period_part}__{original_stem}{file_ext}"
    return unique_path(folder / staged_name)


def check_crewdible_manifest_status(filename, file_path, metadata: CrewdibleUploadMetadata, engine):
    ensure_file_manifest_tables(engine)
    inspection = inspect_crewdible_file(file_path)
    checksum = calculate_sha256(file_path)
    file_size = Path(file_path).stat().st_size

    if not inspection["valid"]:
        return {
            "status": "invalid",
            "rows_in_db": 0,
            "rows_in_file": int(inspection["physical_rows"] or 0),
            "checksum_sha256": checksum,
            "file_size_bytes": file_size,
            "inspection": inspection,
            "table": "staging.file_manifest",
        }

    with engine.connect() as conn:
        same_checksum = conn.execute(text("""
            SELECT rows_detected
            FROM staging.file_manifest
            WHERE source_system = 'crewdible'
              AND data_category = :data_category
              AND checksum_sha256 = :checksum
              AND file_status <> 'invalid'
            ORDER BY uploaded_at DESC
            LIMIT 1
        """), {
            "data_category": metadata.data_category,
            "checksum": checksum,
        }).first()

        same_identity = conn.execute(text("""
            SELECT rows_detected, checksum_sha256
            FROM staging.file_manifest
            WHERE source_system = 'crewdible'
              AND data_category = :data_category
              AND period_year = :period_year
              AND COALESCE(period_month, 0) = COALESCE(:period_month, 0)
              AND original_filename = :filename
              AND file_status <> 'invalid'
            ORDER BY uploaded_at DESC
            LIMIT 1
        """), {
            "data_category": metadata.data_category,
            "period_year": metadata.period_year,
            "period_month": metadata.period_month,
            "filename": filename,
        }).first()

    rows_in_file = int(inspection["physical_rows"] or 0)
    if same_checksum:
        status = "fully_loaded"
        rows_recorded = int(same_checksum.rows_detected or 0)
    elif same_identity:
        rows_recorded = int(same_identity.rows_detected or 0)
        status = "partial" if rows_recorded < rows_in_file else "anomaly"
    else:
        status = "new"
        rows_recorded = 0

    return {
        "status": status,
        "rows_in_db": rows_recorded,
        "rows_in_file": rows_in_file,
        "checksum_sha256": checksum,
        "file_size_bytes": file_size,
        "inspection": inspection,
        "table": "staging.file_manifest",
    }


def stage_crewdible_uploaded_file(uploaded_file, metadata: CrewdibleUploadMetadata, engine):
    """Write an uploaded Crewdible workbook to filesystem staging and manifest."""
    ensure_file_manifest_tables(engine)
    staged_path = build_crewdible_staging_path(uploaded_file.name, metadata)

    with open(staged_path, "wb") as f:
        f.write(uploaded_file.getbuffer())

    inspection = inspect_crewdible_file(staged_path)
    if not inspection["valid"]:
        staged_path.unlink(missing_ok=True)
        missing = ", ".join(inspection.get("missing_headers") or [])
        raise ValueError(f"Format file Crewdible tidak valid. Kolom kurang: {missing}")

    checksum = calculate_sha256(staged_path)
    file_size = staged_path.stat().st_size
    rows_detected = int(inspection["physical_rows"] or 0)

    with engine.begin() as conn:
        manifest_id = conn.execute(text("""
            INSERT INTO staging.file_manifest (
                source_system,
                data_category,
                period_year,
                period_month,
                marketplace,
                fase,
                store_name,
                original_filename,
                staged_filename,
                file_path,
                file_ext,
                file_size_bytes,
                checksum_sha256,
                rows_detected,
                file_status,
                transform_status,
                checked_at
            )
            VALUES (
                'crewdible',
                :data_category,
                :period_year,
                :period_month,
                'crewdible',
                'TRANSACTION',
                'crewdible',
                :original_filename,
                :staged_filename,
                :file_path,
                :file_ext,
                :file_size_bytes,
                :checksum_sha256,
                :rows_detected,
                'staged',
                'pending',
                NOW()
            )
            RETURNING manifest_id
        """), {
            "data_category": metadata.data_category,
            "period_year": metadata.period_year,
            "period_month": metadata.period_month,
            "original_filename": uploaded_file.name,
            "staged_filename": staged_path.name,
            "file_path": str(staged_path),
            "file_ext": staged_path.suffix.lower(),
            "file_size_bytes": file_size,
            "checksum_sha256": checksum,
            "rows_detected": rows_detected,
        }).scalar_one()

    return {
        "manifest_id": int(manifest_id),
        "staged_filename": staged_path.name,
        "file_path": str(staged_path),
        "rows_detected": rows_detected,
        "transaction_rows": int(inspection["transaction_rows"] or 0),
    }
