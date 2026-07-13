"""Shopee file readers for raw marketplace staging tables."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

from scripts.loaders.common import LoadedFrame
from scripts.schema_maps.shopee import (
    SHOPEE_INCOME_ADJ_COLUMN_MAP,
    SHOPEE_INCOME_MAIN_COLUMN_MAP,
    SHOPEE_INCOME_OPF_COLUMN_MAP,
    SHOPEE_INCOME_SD_COLUMN_MAP,
    SHOPEE_INCOME_SF_COLUMN_MAP,
    SHOPEE_ORDER_COLUMN_MAP,
    SHOPEE_REPORT_COLUMN_MAP,
)
from scripts.utils.normalize import prepare_raw_staging_frame


SHOPEE_TABLES = {
    "order": "shopee_orders",
    "report": "shopee_report",
    "income_main": "shopee_income_main",
    "income_order_processing_fee": "shopee_income_order_processing_fee",
    "income_service_fee": "shopee_income_service_fee",
    "income_shipping_discrepancy": "shopee_income_shipping_discrepancy",
    "income_adjustment": "shopee_income_adjustment",
}

TRANSACTION_REPORT_SHEET = "Transaction Report"


def _rows_to_dataframe(rows: list[tuple], header_index: int, data_index: int) -> pd.DataFrame:
    columns = [str(col).strip() if col is not None else None for col in rows[header_index]]
    df = pd.DataFrame(rows[data_index:], columns=columns).astype("string")
    return df.replace("None", pd.NA)


def _find_header_index(rows: list[tuple], expected_columns: set[str], *, minimum_matches: int) -> int | None:
    for index, row in enumerate(rows):
        row_columns = {str(value).strip() for value in row if value is not None}
        if len(row_columns & expected_columns) >= minimum_matches:
            return index
    return None


def _prepare_loaded_frame(
    df: pd.DataFrame,
    *,
    table_name: str,
    column_map: dict[str, str],
    source_path: Path,
    sheet_name: str | None,
) -> LoadedFrame:
    prepared, ignored, missing = prepare_raw_staging_frame(
        df,
        column_map=column_map,
        source_path=source_path,
    )
    return LoadedFrame(
        table_name=table_name,
        dataframe=prepared,
        source_path=source_path,
        sheet_name=sheet_name,
        ignored_columns=ignored,
        missing_columns=missing,
    )


def read_order(path: str | Path) -> LoadedFrame:
    source_path = Path(path)
    if source_path.suffix.lower() == ".csv":
        df = pd.read_csv(source_path, dtype=str)
    else:
        df = pd.read_excel(source_path, dtype=str, engine="openpyxl")
    return _prepare_loaded_frame(
        df,
        table_name=SHOPEE_TABLES["order"],
        column_map=SHOPEE_ORDER_COLUMN_MAP,
        source_path=source_path,
        sheet_name=None,
    )


def read_income(path: str | Path) -> list[LoadedFrame]:
    source_path = Path(path)
    workbook = openpyxl.load_workbook(source_path, data_only=True)
    loaded_frames: list[LoadedFrame] = []

    for sheet_name in workbook.sheetnames:
        sheet_lower = sheet_name.lower()
        rows = list(workbook[sheet_name].values)
        if not rows:
            continue

        if "seller fee" in sheet_lower:
            if len(rows) < 3:
                continue
            df = _rows_to_dataframe(rows, header_index=1, data_index=2)
            columns = list(df.columns)
            if len(columns) > 1 and columns[1] is None:
                columns[1] = "Lihat berdasarkan"
                df.columns = columns

            loaded_frames.append(
                _prepare_loaded_frame(
                    df,
                    table_name=SHOPEE_TABLES["income_order_processing_fee"],
                    column_map=SHOPEE_INCOME_OPF_COLUMN_MAP,
                    source_path=source_path,
                    sheet_name=sheet_name,
                )
            )
            loaded_frames.append(
                _prepare_loaded_frame(
                    df,
                    table_name=SHOPEE_TABLES["income_service_fee"],
                    column_map=SHOPEE_INCOME_SF_COLUMN_MAP,
                    source_path=source_path,
                    sheet_name=sheet_name,
                )
            )

        elif "order processing fee" in sheet_lower:
            df = _rows_to_dataframe(rows, header_index=0, data_index=1)
            loaded_frames.append(
                _prepare_loaded_frame(
                    df,
                    table_name=SHOPEE_TABLES["income_order_processing_fee"],
                    column_map=SHOPEE_INCOME_OPF_COLUMN_MAP,
                    source_path=source_path,
                    sheet_name=sheet_name,
                )
            )

        elif "income" in sheet_lower:
            if len(rows) < 7:
                continue
            df = _rows_to_dataframe(rows, header_index=5, data_index=6)
            if df.empty or len(df.columns) < 5:
                continue
            loaded_frames.append(
                _prepare_loaded_frame(
                    df,
                    table_name=SHOPEE_TABLES["income_main"],
                    column_map=SHOPEE_INCOME_MAIN_COLUMN_MAP,
                    source_path=source_path,
                    sheet_name=sheet_name,
                )
            )

        elif "service fee" in sheet_lower:
            if len(rows) < 3:
                continue
            df = _rows_to_dataframe(rows, header_index=1, data_index=2)
            loaded_frames.append(
                _prepare_loaded_frame(
                    df,
                    table_name=SHOPEE_TABLES["income_service_fee"],
                    column_map=SHOPEE_INCOME_SF_COLUMN_MAP,
                    source_path=source_path,
                    sheet_name=sheet_name,
                )
            )

        elif "shipping fee" in sheet_lower:
            if len(rows) < 3:
                continue
            df = _rows_to_dataframe(rows, header_index=1, data_index=2)
            loaded_frames.append(
                _prepare_loaded_frame(
                    df,
                    table_name=SHOPEE_TABLES["income_shipping_discrepancy"],
                    column_map=SHOPEE_INCOME_SD_COLUMN_MAP,
                    source_path=source_path,
                    sheet_name=sheet_name,
                )
            )

        elif "adjustment" in sheet_lower:
            header_index = None
            for index, row in enumerate(rows):
                first_column = str(row[0]).strip() if row and row[0] is not None else ""
                nearby_columns = " ".join(str(value) for value in row[1:4]).lower()
                if first_column == "No." and ("tanggal" in nearby_columns or "tipe" in nearby_columns):
                    header_index = index
                    break

            if header_index is None:
                continue

            df = _rows_to_dataframe(rows, header_index=header_index, data_index=header_index + 1)
            df = df.loc[:, df.columns.notna()]
            if "No." in df.columns:
                df = df.dropna(subset=["No."])
                df = df[~df["No."].astype(str).str.lower().str.contains("total", na=False)]

            loaded_frames.append(
                _prepare_loaded_frame(
                    df,
                    table_name=SHOPEE_TABLES["income_adjustment"],
                    column_map=SHOPEE_INCOME_ADJ_COLUMN_MAP,
                    source_path=source_path,
                    sheet_name=sheet_name,
                )
            )

    if not loaded_frames:
        raise ValueError(f"No supported Shopee income sheets found in {source_path.name}.")

    return loaded_frames


def read_report(path: str | Path) -> LoadedFrame:
    source_path = Path(path)
    workbook = openpyxl.load_workbook(source_path, data_only=True)
    if TRANSACTION_REPORT_SHEET not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(
            f"Sheet '{TRANSACTION_REPORT_SHEET}' not found in {source_path.name}. "
            f"Available sheets: {available}"
        )

    rows = list(workbook[TRANSACTION_REPORT_SHEET].values)
    header_index = _find_header_index(
        rows,
        set(SHOPEE_REPORT_COLUMN_MAP),
        minimum_matches=5,
    )
    if header_index is None:
        raise ValueError(
            f"Report header not found in sheet '{TRANSACTION_REPORT_SHEET}' "
            f"for {source_path.name}."
        )

    df = _rows_to_dataframe(rows, header_index=header_index, data_index=header_index + 1)
    return _prepare_loaded_frame(
        df,
        table_name=SHOPEE_TABLES["report"],
        column_map=SHOPEE_REPORT_COLUMN_MAP,
        source_path=source_path,
        sheet_name=TRANSACTION_REPORT_SHEET,
    )


def read_file(path: str | Path, phase: str) -> LoadedFrame | list[LoadedFrame]:
    normalized_phase = phase.lower()
    if normalized_phase == "income":
        return read_income(path)
    if normalized_phase in {"order", "orders"}:
        return read_order(path)
    if normalized_phase in {"report", "reports"}:
        return read_report(path)

    raise ValueError(f"Unsupported Shopee phase: {phase}")
