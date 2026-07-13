"""TikTok-Tokopedia file readers for raw marketplace staging tables."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

from scripts.loaders.common import LoadedFrame
from scripts.schema_maps.tiktok_tokopedia import (
    TIKTOK_INCOME_COLUMN_MAP,
    TIKTOK_ORDER_COLUMN_MAP,
    TIKTOK_REPORT_COLUMN_MAP,
)
from scripts.utils.normalize import prepare_raw_staging_frame


TIKTOK_TOKOPEDIA_TABLES = {
    "income": "tiktok_tokopedia_income",
    "order": "tiktok_tokopedia_orders",
    "report": "tiktok_tokopedia_report",
}

ORDER_DETAILS_SHEETS = {"Order details", "Detail pesanan"}
WITHDRAWAL_RECORDS_SHEETS = {"Withdrawal records", "Riwayat penarikan"}


def _active_sheet_dataframe(path: Path, *, skip_description_row: bool = False) -> pd.DataFrame:
    workbook = openpyxl.load_workbook(path, data_only=True)
    rows = list(workbook.active.values)
    if not rows:
        raise ValueError(f"File is empty: {path.name}")

    columns = [str(col).strip() if col is not None else None for col in rows[0]]
    data_rows = rows[2:] if skip_description_row else rows[1:]
    df = pd.DataFrame(data_rows, columns=columns).astype("string")
    return df.replace("None", pd.NA)


def _prepare_loaded_frame(
    df: pd.DataFrame,
    *,
    table_name: str,
    column_map: dict[str, str],
    source_path: Path,
    sheet_name: str | None,
) -> LoadedFrame:
    prepared, ignored, missing = prepare_raw_staging_frame(
        df,
        column_map=column_map,
        source_path=source_path,
    )
    return LoadedFrame(
        table_name=table_name,
        dataframe=prepared,
        source_path=source_path,
        sheet_name=sheet_name,
        ignored_columns=ignored,
        missing_columns=missing,
    )


def read_order(path: str | Path) -> LoadedFrame:
    source_path = Path(path)
    if source_path.suffix.lower() == ".csv":
        df = pd.read_csv(source_path, dtype=str)
    else:
        df = _active_sheet_dataframe(source_path, skip_description_row=True)

    return _prepare_loaded_frame(
        df,
        table_name=TIKTOK_TOKOPEDIA_TABLES["order"],
        column_map=TIKTOK_ORDER_COLUMN_MAP,
        source_path=source_path,
        sheet_name=None,
    )


def read_income(path: str | Path) -> LoadedFrame:
    source_path = Path(path)
    if source_path.suffix.lower() == ".csv":
        df = pd.read_csv(source_path, dtype=str)
        sheet_name = None
    else:
        excel_file = pd.ExcelFile(source_path, engine="openpyxl")
        sheet_name = next(
            (sheet for sheet in excel_file.sheet_names if sheet in ORDER_DETAILS_SHEETS),
            excel_file.sheet_names[0],
        )
        df = pd.read_excel(source_path, sheet_name=sheet_name, dtype=str, engine="openpyxl")

    return _prepare_loaded_frame(
        df,
        table_name=TIKTOK_TOKOPEDIA_TABLES["income"],
        column_map=TIKTOK_INCOME_COLUMN_MAP,
        source_path=source_path,
        sheet_name=sheet_name,
    )


def read_report(path: str | Path) -> LoadedFrame:
    source_path = Path(path)
    if source_path.suffix.lower() == ".csv":
        df = pd.read_csv(source_path, dtype=str).fillna("")
        sheet_name = None
    else:
        workbook = openpyxl.load_workbook(source_path, data_only=True)
        sheet_name = next(
            (sheet for sheet in workbook.sheetnames if sheet in WITHDRAWAL_RECORDS_SHEETS),
            None,
        )
        if sheet_name is None:
            if "income" in str(source_path).lower():
                available = ", ".join(workbook.sheetnames)
                raise ValueError(
                    f"Withdrawal report sheet not found in {source_path.name}. "
                    f"Available sheets: {available}"
                )
            sheet_name = workbook.active.title
        rows = list(workbook[sheet_name].values)
        if not rows:
            raise ValueError(f"Sheet '{sheet_name}' is empty in {source_path.name}.")
        columns = [str(col).strip() if col is not None else None for col in rows[0]]
        df = pd.DataFrame(rows[1:], columns=columns).astype("string").replace("None", pd.NA)

    return _prepare_loaded_frame(
        df,
        table_name=TIKTOK_TOKOPEDIA_TABLES["report"],
        column_map=TIKTOK_REPORT_COLUMN_MAP,
        source_path=source_path,
        sheet_name=sheet_name,
    )


def read_file(path: str | Path, phase: str) -> LoadedFrame:
    normalized_phase = phase.lower()
    if normalized_phase == "income":
        return read_income(path)
    if normalized_phase in {"order", "orders"}:
        return read_order(path)
    if normalized_phase in {"report", "reports"}:
        return read_report(path)

    raise ValueError(f"Unsupported TikTok-Tokopedia phase: {phase}")
