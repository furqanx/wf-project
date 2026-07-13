"""Lazada file readers for raw marketplace staging tables."""

from __future__ import annotations

from pathlib import Path
import openpyxl
import pandas as pd

from scripts.loaders.common import LoadedFrame
from scripts.schema_maps.lazada import (
    LAZADA_INCOME_COLUMN_MAP,
    LAZADA_ORDER_COLUMN_MAP,
    LAZADA_REPORT_COLUMN_MAP,
)
from scripts.utils.normalize import prepare_raw_staging_frame


LAZADA_TABLES = {
    "income": "lazada_income",
    "order": "lazada_orders",
    "report": "lazada_report",
}

BALANCE_TRANSACTIONS_SHEET = "Balance Transactions"


def read_income(path: str | Path) -> LoadedFrame:
    source_path = Path(path)
    df = pd.read_excel(source_path, dtype=str, engine="openpyxl")
    prepared, ignored, missing = prepare_raw_staging_frame(
        df,
        column_map=LAZADA_INCOME_COLUMN_MAP,
        source_path=source_path,
    )
    return LoadedFrame(
        table_name=LAZADA_TABLES["income"],
        dataframe=prepared,
        source_path=source_path,
        sheet_name=None,
        ignored_columns=ignored,
        missing_columns=missing,
    )


def read_order(path: str | Path) -> LoadedFrame:
    source_path = Path(path)
    df = pd.read_excel(source_path, dtype=str, engine="openpyxl")
    prepared, ignored, missing = prepare_raw_staging_frame(
        df,
        column_map=LAZADA_ORDER_COLUMN_MAP,
        source_path=source_path,
    )
    return LoadedFrame(
        table_name=LAZADA_TABLES["order"],
        dataframe=prepared,
        source_path=source_path,
        sheet_name=None,
        ignored_columns=ignored,
        missing_columns=missing,
    )


def read_report(path: str | Path) -> LoadedFrame:
    source_path = Path(path)
    workbook = openpyxl.load_workbook(source_path, data_only=True)
    if BALANCE_TRANSACTIONS_SHEET not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(
            f"Sheet '{BALANCE_TRANSACTIONS_SHEET}' not found in {source_path.name}. "
            f"Available sheets: {available}"
        )

    worksheet = workbook[BALANCE_TRANSACTIONS_SHEET]
    rows = list(worksheet.values)
    if not rows:
        raise ValueError(f"Sheet '{BALANCE_TRANSACTIONS_SHEET}' is empty in {source_path.name}.")

    columns = [str(col).strip() if col is not None else "" for col in rows[0]]
    df = pd.DataFrame(rows[1:], columns=columns).astype("string")
    prepared, ignored, missing = prepare_raw_staging_frame(
        df,
        column_map=LAZADA_REPORT_COLUMN_MAP,
        source_path=source_path,
    )
    return LoadedFrame(
        table_name=LAZADA_TABLES["report"],
        dataframe=prepared,
        source_path=source_path,
        sheet_name=BALANCE_TRANSACTIONS_SHEET,
        ignored_columns=ignored,
        missing_columns=missing,
    )


def read_file(path: str | Path, phase: str) -> LoadedFrame:
    normalized_phase = phase.lower()
    if normalized_phase == "income":
        return read_income(path)
    if normalized_phase in {"order", "orders"}:
        return read_order(path)
    if normalized_phase in {"report", "reports"}:
        return read_report(path)

    raise ValueError(f"Unsupported Lazada phase: {phase}")

