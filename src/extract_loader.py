import pandas as pd
import openpyxl
import os
import re
import zipfile
import xml.etree.ElementTree as ET
from sqlalchemy import text
from src.db_config import logger
from src.notifier import notify_schema_drift, notify_unknown_sheet

def extract_store_name(filename):
    """
    Mendeteksi nama toko dari nama file.
    Contoh: '1. order official 2022.xlsx' -> 'official'
    """
    name_lower = filename.lower()
    
    # Daftar kata kunci nama toko Klien Anda
    # Tambahkan jika ada nama toko baru di masa depan
    store_keywords = [
        'official',
        # Beras Sehat
        'beras sehat',
        'berassehat',
        # Bandar Organik
        'bandar organik',
        'bandar organk',          # typo alias
        'bandarorganik',
        'bandar',                 # alias pendek
        # DIY Jateng
        'diy jateng',
        'wellfarm diy',
        'wellfarmdiyjateng',
        'diy',                    # alias pendek
        # Porice
        'porice',
        # Merapi
        'merapi organik',
        'merapiorganik',
        'merapi',
        # Diabetashop
        'diabetashop',
        'diabetasop',             # typo alias
        'berasdiabetes',          # alias username Shopee
        # Bromo
        'bromo organik',
        'bromoorganik',
        'bromo',
        # Bogor Healthy
        'bogor healthy',
        'bogorhealthy',
        'bogorhealthystore',
        'bogor healhy',           # typo alias
        'bogor',                  # alias pendek
        # Owellness
        'owellness',
        'owellnes',               # typo alias
        # Basecamp
        'basecamp organik',
        'basecamporganik',
        'basecamp',
        # Lembah Organik
        'lembah organik',
        'lembahorganik',
        'lembah',                 # alias pendek
        # Medan Organik
        'medan organik',
        'medanorganik',
        'berasmedanorganik',
        'medan',                  # alias pendek
        # Merbabu Organik
        'merbabu organik',
        'merbabuorganik',
        'merbabu',                # alias pendek
        'merbau',                 # typo alias
        'merababu',               # typo alias
        # Truly Organik
        'truly organik',
        'trulyorganik',
        'truly',                  # alias pendek
        # Diet Healthy
        'diet healthy corner',
        'diet healthy',
        'diethealthy',
        'diet',                   # alias pendek
        # Wellfarm
        'wellfarm id',
        'wellfarm store',
        'wellfarm shop',
        # Pako
        'pako',
        # Solo Organik
        'solo organik',
        'soloorganik',
        'solo',                   # alias pendek
        # Solusi Beras Sehat
        'solusi beras sehat',
        'solusi beras seat',      # typo alias
        'solusiberassehat',
        'solusi',                 # alias pendek
        # Sumber Organik
        'sumber organik shop',
        'sumberorganikshop',
        'sumber organik',
        # Sumber Pangan
        'sumber pangan pokok',
        'sumberpanganpokok',
        'sumber pangan',          # alias pendek
        # Pusat Beras Berkualitas
        'pusat beras berkualitas',
        'pusat beras berkualita', # typo alias
        'pusatberasberkualitas',
        'pusat beras',            # alias pendek
        # Zona Pangan
        'zona pangan',
        'zonapangan',
        'zonapangansehat',
        'zona',                   # alias pendek
        # Porang Sachet
        'porang sachet store',
        'porangsachetstore',
        'porang sachet',
        'porangsachet',
        # Indoporang
        'indoporang market',
        'indoporang',
        # Mekar Organik
        'mekar organik',
        'mekar',                  # alias pendek
        # Organic Groceries
        'organic groceries',
        'organicgroceries',
        'organic',                # alias pendek
        # Mapan Organik
        'mapan organik',
        'mapanorganik',
        'mapan oraganik',         # typo alias
        'mapan',                  # alias pendek
        # Beras Organik ID
        'beras organik id',
        'berasorganikid',
        'organik id',             # alias pendek
        'organikid',              # alias dari filename 'Beras OrganikID' (tanpa spasi)
        # Pundi
        'pundi',
    ]
    
    for store in store_keywords:
        if store in name_lower:
            return store

    return "-"


def extract_tiktok_report_store_name(filename):
    """
    Khusus file REPORT TikTok yang format namanya:
    'Tokopedia [TokpedStore]-TikTok [TikTokStore] [tahun] - REPORT'
    Ekstrak nama toko dari bagian TikTok saja agar tidak salah mapping ke channel Tokopedia.
    """
    name_lower = filename.lower()
    match = re.search(r'tiktok\s+(.+?)\s+\d{4}', name_lower)
    if match:
        tiktok_part = match.group(1).strip()
        result = extract_store_name(tiktok_part)
        if result != '-':
            return result
    return extract_store_name(filename)


def enforce_schema(df, valid_columns, filename, table_name='', file_path='', engine=None):
    """
    Mencegah error Schema Drift. Membuang kolom dari Excel yang tidak ada di tabel Staging SQL.
    Jika ada kolom baru yang tidak dikenal:
      - Log warning
      - Kirim notifikasi Telegram
      - Catat ke staging.schema_drift_log
      - Upload file ke Google Drive
    """
    current_cols   = set(df.columns)
    valid_cols_set = set(valid_columns)

    extra_cols = current_cols - valid_cols_set

    if extra_cols:
        logger.warning(f"⚠️ SCHEMA DRIFT TERDETEKSI pada file {filename}!")
        logger.warning(f"Kolom baru diabaikan agar tidak crash: {extra_cols}")
        df = df.drop(columns=list(extra_cols))
        notify_schema_drift(filename, extra_cols, table_name, file_path, engine)

    return df

# ==========================================
# FUNGSI LOADER KE DATABASE
# ==========================================
def load_to_staging(df, table_name, engine):
    """Menembakkan DataFrame utuh ke tabel Staging PostgreSQL."""
    try:
        # if_exists='append' memastikan data ditambahkan ke bawah, bukan menimpa tabel
        df.to_sql(name=table_name, con=engine, if_exists='append', index=False)
        logger.info(f"✅ Sukses load {len(df)} baris ke tabel {table_name}")
    except Exception as e:
        logger.error(f"❌ Gagal load ke tabel {table_name}. Error: {e}")


def dedup_df(df, subset, table_name, filename):
    """
    Menghapus baris duplikat dari DataFrame berdasarkan kolom kunci unik.
    - Hanya kolom subset yang benar-benar ada di df yang digunakan.
    - NaN/None pada kolom subset diperlakukan sebagai nilai yang sama (perilaku default pandas),
      sehingga dua baris dengan SKU null yang identik tetap terdeteksi sebagai duplikat.
    """
    before = len(df)

    valid_subset = [col for col in subset if col in df.columns]

    if not valid_subset:
        logger.warning(f"⚠️ Tidak ada kolom subset yang valid untuk dedup [{table_name}] pada file {filename}. Dedup dilewati.")
        return df

    df = df.drop_duplicates(subset=valid_subset, keep='first')
    after = len(df)
    dropped = before - after

    if dropped > 0:
        logger.warning(f"🔁 DEDUP [{table_name}] | {filename}: {dropped} baris duplikat dihapus ({before} → {after} baris)")
    else:
        logger.info(f"✔ Tidak ada duplikat ditemukan [{table_name}] | {filename}")

    return df


def filter_existing_rows(df, unique_cols, table_name, engine, filename, global_check=False):
    """
    Menyaring baris yang sudah ada di database.

    - global_check=False (default): query dibatasi WHERE source_filename = filename.
      Mencegah file yang sama di-load ulang. Efisien karena tidak full table scan.
    - global_check=True: query seluruh tabel tanpa filter filename.
      Dipakai saat data dari file berbeda (misal bulanan vs tahunan) bisa tumpang-tindih;
      mencegah duplikasi lintas-file berdasarkan kolom kunci unik saja.
    - NULL pada kolom kunci ditangani dengan sentinel string sebelum anti-join,
      karena pandas merge tidak menganggap NaN == NaN secara default.
    """
    valid_cols = [col for col in unique_cols if col in df.columns]

    if not valid_cols:
        logger.warning(f"⚠️ Tidak ada kolom unik yang valid untuk DB check [{table_name}] | {filename}. Pengecekan DB dilewati.")
        return df

    cols_sql = ', '.join(valid_cols)
    try:
        with engine.connect() as conn:
            if global_check:
                result = conn.execute(text(f"SELECT {cols_sql} FROM {table_name}"))
            else:
                result = conn.execute(
                    text(f"SELECT {cols_sql} FROM {table_name} WHERE source_filename = :fn"),
                    {"fn": filename}
                )
            existing_df = pd.DataFrame(result.fetchall(), columns=valid_cols)
    except Exception as e:
        logger.error(f"❌ Gagal query DB untuk filter_existing_rows [{table_name}] | {filename}: {e}")
        return df

    if existing_df.empty:
        # File belum pernah di-load sama sekali, tidak perlu filter
        return df

    before = len(df)

    # Gunakan sentinel untuk menangani NULL agar anti-join bekerja benar
    SENTINEL = '__NULL__'
    df_check    = df[valid_cols].copy().fillna(SENTINEL)
    exist_check = existing_df.fillna(SENTINEL)

    merged = df_check.merge(exist_check, on=valid_cols, how='left', indicator=True)
    mask   = (merged['_merge'] == 'left_only').values
    df     = df[mask]

    after   = len(df)
    skipped = before - after

    if skipped == before:
        logger.info(f"⏭ SKIP [{table_name}] | {filename}: semua {before} baris sudah ada di DB, file dilewati")
    elif skipped > 0:
        logger.warning(f"🔍 DB CHECK [{table_name}] | {filename}: {skipped} baris sudah ada di DB, {after} baris baru akan di-load")

    return df


# ==========================================
# KONSTANTA KOLOM VALID (SCHEMA GUARD)
# ==========================================
# Ini adalah daftar kolom yang diizinkan masuk ke tabel Staging ORDER.
# Ditambah dengan kolom metadata 'nama_toko' dan 'source_filename'

VALID_TIKTOK_ORDER_COLS = [
    'order_id', 'order_status', 'order_substatus', 'cancelation_return_type', 'normal_or_pre_order', 'sku_id', 'seller_sku', 'product_name', 'variation', 'quantity', 'sku_quantity_of_return', 'sku_unit_original_price', 'sku_subtotal_before_discount', 'sku_platform_discount', 'sku_seller_discount', 'sku_subtotal_after_discount', 'shipping_fee_after_discount', 'original_shipping_fee', 'shipping_fee_seller_discount', 'shipping_fee_platform_discount', 'distance_shipping_fee', 'distance_fee', 'order_refund_amount', 'payment_platform_discount', 'buyer_service_fee', 'handling_fee', 'shipping_insurance', 'item_insurance', 'order_amount', 'created_time', 'paid_time', 'rts_time', 'shipped_time', 'delivered_time', 'cancelled_time', 'cancel_by', 'cancel_reason', 'fulfillment_type', 'warehouse_name', 'tracking_id', 'delivery_option', 'shipping_provider_name', 'buyer_message', 'buyer_username', 'recipient', 'phone_number', 'zipcode', 'country', 'province', 'regency_and_city', 'districts', 'villages', 'detail_address', 'additional_address_information', 'payment_method', 'weight_kg', 'product_category', 'package_id', 'purchase_channel', 'seller_note', 'checked_status', 'checked_marked_by', 'tokopedia_invoice_number',
    'nama_toko', 'source_filename' # Metadata
]

VALID_SHOPEE_ORDER_COLS = [
    'no_pesanan', 'status_pesanan', 'alasan_pembatalan', 'status_pembatalan_pengembalian', 'no_resi', 'opsi_pengiriman', 'antar_ke_counter_pickup', 'pesanan_harus_dikirimkan_sebelum', 'waktu_pengiriman_diatur', 'waktu_pesanan_dibuat', 'waktu_pembayaran_dilakukan', 'metode_pembayaran', 'sku_induk', 'nama_produk', 'nomor_referensi_sku', 'nama_variasi', 'harga_awal', 'harga_setelah_diskon', 'jumlah', 'returned_quantity', 'total_harga_produk', 'total_diskon', 'diskon_dari_penjual', 'diskon_dari_shopee', 'berat_produk', 'jumlah_produk_di_pesan', 'total_berat', 'nama_gudang', 'voucher_ditanggung_penjual', 'cashback_koin', 'voucher_ditanggung_shopee', 'paket_diskon', 'paket_diskon_dari_shopee', 'paket_diskon_dari_penjual', 'potongan_koin_shopee', 'diskon_kartu_kredit', 'ongkos_kirim_dibayar_oleh_pembeli', 'estimasi_potongan_biaya_pengiriman', 'ongkos_kirim_pengembalian_barang', 'total_pembayaran', 'perkiraan_ongkos_kirim', 'catatan_dari_pembeli', 'catatan', 'username_pembeli', 'nama_penerima', 'no_telepon', 'alamat_pengiriman', 'kota_kabupaten', 'provinsi', 'waktu_pesanan_selesai',
    'nama_toko', 'source_filename' # Metadata
]

VALID_LAZADA_ORDER_COLS = [
    'order_item_id', 'order_type', 'guarantee', 'delivery_type', 'lazada_id', 'seller_sku', 'lazada_sku', 'warehouse', 'create_time', 'update_time', 'rts_sla', 'tts_sla', 'order_number', 'invoice_required', 'invoice_number', 'delivered_date', 'customer_name', 'customer_email', 'national_registration_number', 'shipping_name', 'shipping_address', 'shipping_address2', 'shipping_address3', 'shipping_address4', 'shipping_address5', 'shipping_phone', 'shipping_phone2', 'shipping_city', 'shipping_post_code', 'shipping_country', 'shipping_region', 'billing_name', 'billing_addr', 'billing_addr2', 'billing_addr3', 'billing_addr4', 'billing_addr5', 'billing_phone', 'billing_phone2', 'billing_city', 'billing_post_code', 'billing_country', 'tax_code', 'branch_number', 'tax_invoice_requested', 'pay_method', 'paid_price', 'unit_price', 'seller_discount_total', 'shipping_fee', 'wallet_credit', 'item_name', 'variation', 'cd_shipping_provider', 'shipping_provider', 'shipment_type_name', 'shipping_provider_type', 'cd_tracking_code', 'tracking_code', 'tracking_url', 'shipping_provider_fm', 'tracking_code_fm', 'tracking_url_fm', 'promised_shipping_time', 'premium', 'status', 'buyer_failed_delivery_return_initiator', 'buyer_failed_delivery_reason', 'buyer_failed_delivery_detail', 'buyer_failed_delivery_user_name', 'bundle_id', 'semi_managed', 'flexible_delivery_time', 'bundle_discount', 'refund_amount', 'seller_note',
    'nama_toko', 'source_filename' # Metadata
]


VALID_TIKTOK_INCOME_COLS = [
    'order_adjustment_id', 'type', 'order_created_time', 'order_settled_time', 'currency', 'total_settlement_amount', 'total_revenue', 'subtotal_after_seller_discounts', 'subtotal_before_discounts', 'seller_discounts', 'distance_item_fee_from_horizon_plus_program', 'refund_subtotal_after_seller_discounts', 'refund_subtotal_before_seller_discounts', 'refund_of_seller_discounts', 'total_fees', 'platform_commission_fee', 'pre_order_service_fee', 'mall_service_fee', 'payment_fee', 'shipping_cost', 'shipping_costs_passed_on_to_the_logistics_provider', 'replacement_shipping_fee_passed_on_to_the_customer', 'exchange_shipping_fee_passed_on_to_the_customer', 'shipping_cost_borne_by_the_platform', 'shipping_cost_paid_by_the_customer', 'refunded_shipping_cost_paid_by_the_customer', 'return_shipping_costs_passed_on_to_the_customer', 'shipping_cost_subsidy', 'distance_shipping_fee_from_horizon_plus_program', 'affiliate_commission', 'affiliate_partner_commission', 'affiliate_shop_ads_commission', 'affiliate_partner_shop_ads_commission', 'shipping_fee_program_service_fee', 'dynamic_commission', 'bonus_cashback_service_fee', 'live_specials_service_fee', 'voucher_xtra_service_fee', 'order_processing_fee', 'eams_program_service_fee', 'brands_crazy_deals_flash_sale_service_fee', 'dilayani_tokopedia_fee', 'dilayani_tokopedia_handling_fee', 'paylater_program_fee', 'campaign_resource_fee', 'installation_service_fee', 'article_22_income_tax_withheld', 'platform_special_service_fee', 'gmv_max_ad_fee', 'gmv_max_coupon', 'ajustment_amount', 'related_order_id', 'shipping_fee_adjustment', 'shipping_fee_compensation', 'chargeback', 'customer_service_compensation', 'promotion_adjustment', 'platform_compensation', 'platform_penalty', 'sample_shipping_fee', 'logistics_reimbursement', 'platform_reimbursement', 'deductions_incurred_by_seller', 'shipping_fee_rebate', 'warehouse_service_fee', 'platform_commission_adjustment', 'platform_commission_compensation', 'transaction_fee_adjustment', 'top_up_for_ads_from_settled_balances', 'campaign_package', 'additional_campaign_package', 'gmv_payment_for_tiktok_ads', 'gmv_payment_for_promote', 'shipping_insurance_compensation', 'other_adjustment', 'customer_payment', 'customer_refund', 'seller_co_funded_voucher_discount', 'refund_of_seller_co_funded_voucher_discount', 'platform_discounts', 'refund_of_platform_discounts', 'platform_co_funded_voucher_discounts', 'refund_of_platform_co_funded_voucher_discounts', 'seller_shipping_cost_discount', 'estimated_package_weight', 'actual_package_weight', 'order_source', 'shopping_center_items',
    'nama_toko', 'source_filename' # Metadata
]

VALID_SHOPEE_INCOME_OPF_COLS = [
    'no', 'lihat_berdasarkan', 'no_pesanan', 'id_produk', 'nama_produk', 'biaya_proses_pesanan', 'biaya_proses_pesanan_per_produk_prorata',   # <--- JANGAN LUPA KOMA DI SINI
    'nama_toko', 'source_filename' # Metadata
]

VALID_SHOPEE_INCOME_MAIN_COLS = [
    'no', 'no_pesanan', 'no_pengajuan', 'username_pembeli', 'waktu_pesanan_dibuat', 'metode_pembayaran_pembeli', 'tanggal_dana_dilepaskan', 'harga_asli_produk', 'total_diskon_produk', 'jumlah_pengembalian_dana_ke_pembeli', 'diskon_produk_dari_shopee', 'voucher_disponsor_oleh_penjual', 'voucher_co_fund_disponsor_oleh_penjual', 'cashback_koin_disponsori_penjual', 'cashback_koin_co_fund_disponsori_penjual', 'ongkir_dibayar_pembeli', 'diskon_ongkir_ditanggung_jasa_kirim', 'gratis_ongkir_dari_shopee', 'ongkir_yang_diteruskan_oleh_shopee_ke_jasa_kirim', 'ongkos_kirim_pengembalian_barang', 'kembali_ke_biaya_pengiriman_pengirim', 'pengembalian_biaya_kirim', 'biaya_komisi_ams', 'biaya_administrasi_termasuk_ppn_11', 'biaya_layanan', 'biaya_proses_pesanan', 'premi', 'biaya_program_hemat_biaya_kirim', 'biaya_transaksi', 'biaya_kampanye', 'bea_masuk_ppn_pph', 'biaya_isi_saldo_otomatis_dari_penghasilan', 'total_penghasilan', 'kode_voucher', 'kompensasi', 'promo_gratis_ongkir_dari_penjual', 'jasa_kirim', 'nama_kurir', 'pengembalian_dana_ke_pembeli', 'pro_rata_koin_yang_ditukarkan_untuk_pengembalian_barang', 'pro_rata_voucher_shopee_untuk_pengembalian_barang', 'pro_rated_bank_payment_channel_promotion_for_return_refund_item', 'pro_rated_shopee_payment_channel_promotion_for_return_refund_it',
    'nama_toko', 'source_filename' # Metadata
]

VALID_SHOPEE_INCOME_SF_COLS = [
    'no', 'no_pesanan',
    'biaya_layanan_cashback_xtra', 'biaya_layanan_cashbackxtra', 'biaya_layanan_gratis_ongkir_xtra', 'biaya_layanan_gratis_ongkir_xtra_2',
    'biaya_layanan_promo_xtra', 'biaya_pembayaran', 'biaya_program_shopee_live_xtra',
    'biaya_campaign_1_1', 'biaya_campaign_2_2', 'biaya_campaign_3_3',
    'biaya_campaign_4_4', 'biaya_campaign_5_5', 'biaya_campaign_6_6',
    'biaya_campaign_7_7', 'biaya_campaign_8_8', 'biaya_campaign_9_9',
    'biaya_campaign_10_10', 'biaya_campaign_11_11', 'biaya_campaign_12_12',
    'nama_toko', 'source_filename' # Metadata
]

VALID_SHOPEE_INCOME_SD_COLS = [
    'no_pesanan', 'estimasi_ongkos_kirim', 'ongkos_kirim_yang_dibayarkan_ke_jasa_kirim', 'discrepancy_reason',
    'nama_toko', 'source_filename' # Metadata
]

VALID_SHOPEE_INCOME_ADJ_COLS = [
    'no', 'tanggal_penyesuaian_dibuat', 'tipe_penyesuaian_deskripsi', 'alasan_penyesuaian', 'biaya_penyesuaian', 'no_pesanan_terhubung', 'tanggal_dana_dilepaskan',
    'nama_toko', 'source_filename' # Metadata
]

VALID_LAZADA_INCOME_COLS = [
    'periode_laporan', 'nomor_laporan', 'tanggal_transaksi', 'nama_biaya', 'jumlah_termasuk_pajak', 'vat_amount', 'status_pelepasan_dana', 'tanggal_dilepas', 'komentar', 'tanggal_pesanan_dibuat', 'nomor_pesanan', 'id_pesanan', 'sku_penjual', 'lazada_sku', 'wht_amount', 'wht_termasuk_dalam_jumlah', 'status_pesanan', 'nama_produk', 'short_code',
    'nama_toko', 'source_filename' # Metadata
]


VALID_TIKTOK_REPORT_COLS = [
    'type', 'reference_id', 'request_time', 'amount', 'status', 'success_time', 'bank_account',
    'nama_toko', 'source_filename' # Metadata
]

VALID_SHOPEE_REPORT_COLS = [
    'tanggal_transaksi', 'tipe_transaksi', 'deskripsi', 'no_pesanan',
    'jenis_transaksi', 'jumlah', 'status', 'saldo_akhir',
    'nama_toko', 'source_filename' # Metadata
]

VALID_LAZADA_REPORT_COLS = [
    'transaction_number', 'transaction_time', 'type', 'sub_type', 'amount', 'remarks',
    'nama_toko', 'source_filename' # Metadata
]



# ==========================================
# DICTIONARY MAPPING KOLOM (EXCEL -> SQL)
# ==========================================
TIKTOK_ORDER_COLUMN_MAP = {
    'Order ID': 'order_id',
    'Order Status': 'order_status',
    'Order Substatus': 'order_substatus',
    'Cancelation/Return Type': 'cancelation_return_type',
    'Normal or Pre-order': 'normal_or_pre_order',
    'SKU ID': 'sku_id',
    'Seller SKU': 'seller_sku',
    'Product Name': 'product_name',
    'Variation': 'variation',
    'Quantity': 'quantity',
    'Sku Quantity of return': 'sku_quantity_of_return',
    'SKU Unit Original Price': 'sku_unit_original_price',
    'SKU Subtotal Before Discount': 'sku_subtotal_before_discount',
    'SKU Platform Discount': 'sku_platform_discount',
    'SKU Seller Discount': 'sku_seller_discount',
    'SKU Subtotal After Discount': 'sku_subtotal_after_discount',
    'Shipping Fee After Discount': 'shipping_fee_after_discount',
    'Original Shipping Fee': 'original_shipping_fee',
    'Shipping Fee Seller Discount': 'shipping_fee_seller_discount',
    'Shipping Fee Platform Discount': 'shipping_fee_platform_discount',
    'Distance Shipping Fee': 'distance_shipping_fee',
    'Distance Fee': 'distance_fee',
    'Order Refund Amount': 'order_refund_amount',
    'Payment platform discount': 'payment_platform_discount',
    'Buyer Service Fee': 'buyer_service_fee',
    'Handling Fee': 'handling_fee',
    'Shipping Insurance': 'shipping_insurance',
    'Item Insurance': 'item_insurance',
    'Order Amount': 'order_amount',
    'Created Time': 'created_time',
    'Paid Time': 'paid_time',
    'RTS Time': 'rts_time',
    'Shipped Time': 'shipped_time',
    'Delivered Time': 'delivered_time',
    'Cancelled Time': 'cancelled_time',
    'Cancel By': 'cancel_by',
    'Cancel Reason': 'cancel_reason',
    'Fulfillment Type': 'fulfillment_type',
    'Warehouse Name': 'warehouse_name',
    'Tracking ID': 'tracking_id',
    'Delivery Option': 'delivery_option',
    'Shipping Provider Name': 'shipping_provider_name',
    'Buyer Message': 'buyer_message',
    'Buyer Username': 'buyer_username',
    'Recipient': 'recipient',
    'Phone #': 'phone_number',
    'Zipcode': 'zipcode',
    'Country': 'country',
    'Province': 'province',
    'Regency and City': 'regency_and_city',
    'Districts': 'districts',
    'Villages': 'villages',
    'Detail Address': 'detail_address',
    'Additional address information': 'additional_address_information',
    'Payment Method': 'payment_method',
    'Weight(kg)': 'weight_kg',
    'Product Category': 'product_category',
    'Package ID': 'package_id',
    'Purchase Channel': 'purchase_channel',
    'Seller Note': 'seller_note',
    'Checked Status': 'checked_status',
    'Checked Marked by': 'checked_marked_by',
    'Tokopedia Invoice Number': 'tokopedia_invoice_number'
}

SHOPEE_ORDER_COLUMN_MAP = {
    'No. Pesanan': 'no_pesanan',
    'Status Pesanan': 'status_pesanan',
    'Alasan Pembatalan': 'alasan_pembatalan',
    'Status Pembatalan/ Pengembalian': 'status_pembatalan_pengembalian',
    'No. Resi': 'no_resi',
    'Opsi Pengiriman': 'opsi_pengiriman',
    'Antar ke counter/ pick-up': 'antar_ke_counter_pickup',
    'Pesanan Harus Dikirimkan Sebelum (Menghindari keterlambatan)': 'pesanan_harus_dikirimkan_sebelum',
    'Waktu Pengiriman Diatur': 'waktu_pengiriman_diatur',
    'Waktu Pesanan Dibuat': 'waktu_pesanan_dibuat',
    'Waktu Pembayaran Dilakukan': 'waktu_pembayaran_dilakukan',
    'Metode Pembayaran': 'metode_pembayaran',
    'SKU Induk': 'sku_induk',
    'Nama Produk': 'nama_produk',
    'Nomor Referensi SKU': 'nomor_referensi_sku',
    'Nama Variasi': 'nama_variasi',
    'Harga Awal': 'harga_awal',
    'Harga Setelah Diskon': 'harga_setelah_diskon',
    'Jumlah': 'jumlah',
    'Returned quantity': 'returned_quantity',
    'Total Harga Produk': 'total_harga_produk',
    'Dibayar Pembeli': 'total_harga_produk',      # alias
    'Total Diskon': 'total_diskon',
    'Diskon Dari Penjual': 'diskon_dari_penjual',
    'Diskon Dari Shopee': 'diskon_dari_shopee',
    'Berat Produk': 'berat_produk',
    'Jumlah Produk di Pesan': 'jumlah_produk_di_pesan',
    'Total Berat': 'total_berat',
    'Nama Gudang' : 'nama_gudang',  ############
    'Voucher Ditanggung Penjual': 'voucher_ditanggung_penjual',
    'Cashback Koin': 'cashback_koin',
    'Voucher Ditanggung Shopee': 'voucher_ditanggung_shopee',
    'Paket Diskon': 'paket_diskon',
    'Paket Diskon (Diskon dari Shopee)': 'paket_diskon_dari_shopee',
    'Paket Diskon (Diskon dari Penjual)': 'paket_diskon_dari_penjual',
    'Potongan Koin Shopee': 'potongan_koin_shopee',
    'Diskon Kartu Kredit': 'diskon_kartu_kredit',
    'Ongkos Kirim Dibayar oleh Pembeli': 'ongkos_kirim_dibayar_oleh_pembeli',
    'Estimasi Potongan Biaya Pengiriman': 'estimasi_potongan_biaya_pengiriman',
    'Ongkos Kirim Pengembalian Barang': 'ongkos_kirim_pengembalian_barang',
    'Total Pembayaran': 'total_pembayaran',
    'Perkiraan Ongkos Kirim': 'perkiraan_ongkos_kirim',
    'Catatan dari Pembeli': 'catatan_dari_pembeli',
    'Catatan': 'catatan',
    'Username (Pembeli)': 'username_pembeli',
    'Nama Penerima': 'nama_penerima',
    'No. Telepon': 'no_telepon',
    'Alamat Pengiriman': 'alamat_pengiriman',
    'Kota/Kabupaten': 'kota_kabupaten',
    'Provinsi': 'provinsi',
    'Waktu Pesanan Selesai': 'waktu_pesanan_selesai'
}

LAZADA_ORDER_COLUMN_MAP = {
    'orderItemId': 'order_item_id',
    'orderType': 'order_type',
    'Guarantee': 'guarantee',
    'deliveryType': 'delivery_type',
    'lazadaId': 'lazada_id',
    'sellerSku': 'seller_sku',
    'lazadaSku': 'lazada_sku',
    'wareHouse': 'warehouse',
    'createTime': 'create_time',
    'updateTime': 'update_time',
    'rtsSla': 'rts_sla',
    'ttsSla': 'tts_sla',
    'orderNumber': 'order_number',
    'invoiceRequired': 'invoice_required',
    'invoiceNumber': 'invoice_number',
    'deliveredDate': 'delivered_date',
    'customerName': 'customer_name',
    'customerEmail': 'customer_email',
    'nationalRegistrationNumber': 'national_registration_number',
    'shippingName': 'shipping_name',
    'shippingAddress': 'shipping_address',
    'shippingAddress2': 'shipping_address2',
    'shippingAddress3': 'shipping_address3',
    'shippingAddress4': 'shipping_address4',
    'shippingAddress5': 'shipping_address5',
    'shippingPhone': 'shipping_phone',
    'shippingPhone2': 'shipping_phone2',
    'shippingCity': 'shipping_city',
    'shippingPostCode': 'shipping_post_code',
    'shippingCountry': 'shipping_country',
    'shippingRegion': 'shipping_region',
    'billingName': 'billing_name',
    'billingAddr': 'billing_addr',
    'billingAddr2': 'billing_addr2',
    'billingAddr3': 'billing_addr3',
    'billingAddr4': 'billing_addr4',
    'billingAddr5': 'billing_addr5',
    'billingPhone': 'billing_phone',
    'billingPhone2': 'billing_phone2',
    'billingCity': 'billing_city',
    'billingPostCode': 'billing_post_code',
    'billingCountry': 'billing_country',
    'taxCode': 'tax_code',
    'branchNumber': 'branch_number',
    'taxInvoiceRequested': 'tax_invoice_requested',
    'payMethod': 'pay_method',
    'paidPrice': 'paid_price',
    'unitPrice': 'unit_price',
    'sellerDiscountTotal': 'seller_discount_total',
    'shippingFee': 'shipping_fee',
    'walletCredit': 'wallet_credit',
    'itemName': 'item_name',
    'variation': 'variation',
    'cdShippingProvider': 'cd_shipping_provider',
    'shippingProvider': 'shipping_provider',
    'shipmentTypeName': 'shipment_type_name',
    'shippingProviderType': 'shipping_provider_type',
    'cdTrackingCode': 'cd_tracking_code',
    'trackingCode': 'tracking_code',
    'trackingUrl': 'tracking_url',
    'shippingProviderFM': 'shipping_provider_fm',
    'trackingCodeFM': 'tracking_code_fm',
    'trackingUrlFM': 'tracking_url_fm',
    'promisedShippingTime': 'promised_shipping_time',
    'premium': 'premium',
    'status': 'status',
    'buyerFailedDeliveryReturnInitiator': 'buyer_failed_delivery_return_initiator',
    'buyerFailedDeliveryReason': 'buyer_failed_delivery_reason',
    'buyerFailedDeliveryDetail': 'buyer_failed_delivery_detail',
    'buyerFailedDeliveryUserName': 'buyer_failed_delivery_user_name',
    'bundleId': 'bundle_id',
    'semiManaged': 'semi_managed',
    'flexibleDeliveryTime': 'flexible_delivery_time',
    'bundleDiscount': 'bundle_discount',
    'refundAmount': 'refund_amount',
    'sellerNote': 'seller_note'
}


TIKTOK_INCOME_COLUMN_MAP = {
    'Order/adjustment ID': 'order_adjustment_id', 'Type': 'type', 'Order created time': 'order_created_time', 'Order settled time': 'order_settled_time', 'Currency': 'currency', 'Total settlement amount': 'total_settlement_amount', 'Total Revenue': 'total_revenue', 'Subtotal after seller discounts': 'subtotal_after_seller_discounts', 'Subtotal before discounts': 'subtotal_before_discounts', 'Seller discounts': 'seller_discounts', 'Distance item fee from Horizon+ Program': 'distance_item_fee_from_horizon_plus_program', 'Refund subtotal after seller discounts': 'refund_subtotal_after_seller_discounts', 'Refund subtotal before seller discounts': 'refund_subtotal_before_seller_discounts', 'Refund of seller discounts': 'refund_of_seller_discounts', 'Total Fees': 'total_fees', 'Platform commission fee': 'platform_commission_fee', 'Pre-order service fee': 'pre_order_service_fee', 'Mall service fee': 'mall_service_fee', 'Payment Fee': 'payment_fee', 'Shipping cost': 'shipping_cost', 'Shipping costs passed on to the logistics provider': 'shipping_costs_passed_on_to_the_logistics_provider', 'Replacement shipping fee (passed on to the customer)': 'replacement_shipping_fee_passed_on_to_the_customer', 'Exchange shipping fee (passed on to the customer)': 'exchange_shipping_fee_passed_on_to_the_customer', 'Shipping cost borne by the platform': 'shipping_cost_borne_by_the_platform', 'Shipping cost paid by the customer': 'shipping_cost_paid_by_the_customer', 'Refunded shipping cost paid by the customer': 'refunded_shipping_cost_paid_by_the_customer', 'Return shipping costs (passed on to the customer)': 'return_shipping_costs_passed_on_to_the_customer', 'Shipping cost subsidy': 'shipping_cost_subsidy', 'Distance shipping fee from Horizon+ Program': 'distance_shipping_fee_from_horizon_plus_program', 'Affiliate Commission': 'affiliate_commission', 'Affiliate partner commission': 'affiliate_partner_commission', 'Affiliate Shop Ads commission': 'affiliate_shop_ads_commission', 'Affiliate Partner shop ads commission': 'affiliate_partner_shop_ads_commission', 'Shipping Fee Program service fee': 'shipping_fee_program_service_fee', 'Dynamic commission': 'dynamic_commission', 'Bonus cashback service fee': 'bonus_cashback_service_fee', 'LIVE Specials service fee': 'live_specials_service_fee', 'Voucher Xtra service fee': 'voucher_xtra_service_fee', 'Order processing fee': 'order_processing_fee', 'EAMS Program service fee': 'eams_program_service_fee', 'Brands Crazy Deals/Flash Sale service fee': 'brands_crazy_deals_flash_sale_service_fee', 'Dilayani Tokopedia fee': 'dilayani_tokopedia_fee', 'Dilayani Tokopedia handling fee': 'dilayani_tokopedia_handling_fee', 'PayLater program fee': 'paylater_program_fee', 'Campaign resource fee': 'campaign_resource_fee', 'Installation service fee': 'installation_service_fee', 'Article 22 Income Tax withheld': 'article_22_income_tax_withheld', 'Platform special service fee': 'platform_special_service_fee', 'GMV Max ad fee': 'gmv_max_ad_fee', 'GMV Max Coupon': 'gmv_max_coupon', 'Ajustment amount': 'ajustment_amount', 'Related order ID': 'related_order_id', 
    'Shipping fee adjustment': 'shipping_fee_adjustment', 'Shipping fee compensation': 'shipping_fee_compensation', 'Chargeback': 'chargeback', 'Customer service compensation': 'customer_service_compensation', 'Promotion adjustment': 'promotion_adjustment', 'Platform compensation': 'platform_compensation', 'Platform penalty': 'platform_penalty', 'Sample shipping fee': 'sample_shipping_fee', 'Logistics reimbursement': 'logistics_reimbursement', 'Platform reimbursement': 'platform_reimbursement', 'Deductions incurred by seller': 'deductions_incurred_by_seller', 'Shipping fee rebate': 'shipping_fee_rebate', 'Warehouse service fee': 'warehouse_service_fee', 'Platform commission adjustment': 'platform_commission_adjustment', 'Platform commission compensation': 'platform_commission_compensation', 'Transaction fee adjustment': 'transaction_fee_adjustment', 'Top up for ads from settled balances': 'top_up_for_ads_from_settled_balances', 'Campaign Package': 'campaign_package', 'Additional Campaign Package': 'additional_campaign_package', 'GMV Payment for TikTok Ads': 'gmv_payment_for_tiktok_ads', 'GMV Payment for Promote': 'gmv_payment_for_promote', 'Shipping insurance compensation': 'shipping_insurance_compensation', 'Other adjustment': 'other_adjustment', 
    'Customer payment': 'customer_payment', 
    'Customer refund': 'customer_refund', 
    'Seller co-funded voucher discount': 'seller_co_funded_voucher_discount', 
    'Refund of seller co-funded voucher discount': 'refund_of_seller_co_funded_voucher_discount', 
    'Platform discounts': 'platform_discounts', 
    'Refund of platform discounts': 'refund_of_platform_discounts', 'Platform co-funded voucher discounts': 'platform_co_funded_voucher_discounts', 
    'Refund of platform co-funded voucher discounts': 'refund_of_platform_co_funded_voucher_discounts', 
    'Seller shipping cost discount': 'seller_shipping_cost_discount',
    'Estimated package weight (g)': 'estimated_package_weight',
    'Actual package weight (g)': 'actual_package_weight',
    'Order Source': 'order_source',
    'Shopping center items': 'shopping_center_items'
}

SHOPEE_INCOME_OPF_COLUMN_MAP = {
    'No.': 'no', 
    'Lihat berdasarkan': 'lihat_berdasarkan',
    # 'No. Perusahaan': 'no_perusahaan', #
    'No. Pesanan': 'no_pesanan',
    'ID Produk': 'id_produk', 
    'Nama Produk': 'nama_produk', 
    'Biaya Proses Pesanan': 'biaya_proses_pesanan', 
    # 'Biaya Proses Pesanan per Produk (Prorata)': 'biaya_proses_pesanan_per_produk_prorata' #
    'Biaya Proses Pesanan per Produk (Prorata harga produk tiap pesanan)': 'biaya_proses_pesanan_per_produk_prorata'
}

SHOPEE_INCOME_MAIN_COLUMN_MAP = {
    'No.': 'no',
    'No. Pesanan': 'no_pesanan',
    'No. Pengajuan': 'no_pengajuan',
    'Username (Pembeli)': 'username_pembeli',
    'Waktu Pesanan Dibuat': 'waktu_pesanan_dibuat',
    'Metode pembayaran pembeli': 'metode_pembayaran_pembeli',
    'Tanggal Dana Dilepaskan': 'tanggal_dana_dilepaskan',
    'Harga Asli Produk': 'harga_asli_produk',
    'Total Diskon Produk': 'total_diskon_produk',
    'Jumlah Pengembalian Dana ke Pembeli': 'jumlah_pengembalian_dana_ke_pembeli',
    'Diskon Produk dari Shopee': 'diskon_produk_dari_shopee',
    'Voucher disponsor oleh Penjual': 'voucher_disponsor_oleh_penjual',
    'Voucher co-fund disponsor oleh Penjual': 'voucher_co_fund_disponsor_oleh_penjual',
    'Cashback Koin disponsori Penjual': 'cashback_koin_disponsori_penjual',
    'Cashback Koin Co-fund disponsori Penjual': 'cashback_koin_co_fund_disponsori_penjual',
    'Ongkir Dibayar Pembeli': 'ongkir_dibayar_pembeli',
    'Diskon Ongkir Ditanggung Jasa Kirim': 'diskon_ongkir_ditanggung_jasa_kirim',
    'Gratis Ongkir dari Shopee': 'gratis_ongkir_dari_shopee',
    'Ongkir yang Diteruskan oleh Shopee ke Jasa Kirim': 'ongkir_yang_diteruskan_oleh_shopee_ke_jasa_kirim',
    'Ongkos Kirim Pengembalian Barang': 'ongkos_kirim_pengembalian_barang',
    'Kembali ke Biaya Pengiriman Pengirim': 'kembali_ke_biaya_pengiriman_pengirim',
    'Pengembalian Biaya Kirim': 'pengembalian_biaya_kirim',
    'Biaya Komisi AMS': 'biaya_komisi_ams',
    'Biaya Administrasi (termasuk PPN 11%)': 'biaya_administrasi_termasuk_ppn_11',
    'Biaya Administrasi': 'biaya_administrasi_termasuk_ppn_11',              # alias 2023
    'Biaya Layanan': 'biaya_layanan',
    'Biaya Proses Pesanan': 'biaya_proses_pesanan',
    'Premi': 'premi',
    'Biaya Program Hemat Biaya Kirim': 'biaya_program_hemat_biaya_kirim',
    'Biaya Transaksi': 'biaya_transaksi',
    'Biaya Kampanye': 'biaya_kampanye',
    'Bea Masuk, PPN & PPh': 'bea_masuk_ppn_pph',
    'Biaya Isi Saldo Otomatis (dari Penghasilan)': 'biaya_isi_saldo_otomatis_dari_penghasilan',
    'Total Penghasilan': 'total_penghasilan',
    'Kode Voucher': 'kode_voucher',
    'Kompensasi': 'kompensasi',
    'Promo Gratis Ongkir dari Penjual': 'promo_gratis_ongkir_dari_penjual',
    'Jasa Kirim': 'jasa_kirim',
    'Nama Kurir': 'nama_kurir',
    'Pengembalian Dana ke Pembeli': 'pengembalian_dana_ke_pembeli',
    'Pro-rata Koin yang Ditukarkan untuk Pengembalian Barang': 'pro_rata_koin_yang_ditukarkan_untuk_pengembalian_barang',
    'Pro-rata Voucher Shopee untuk Pengembalian Barang': 'pro_rata_voucher_shopee_untuk_pengembalian_barang',
    'Pro-rated Bank Payment Channel Promotion for return refund Items': 'pro_rated_bank_payment_channel_promotion_for_return_refund_item',
    'Pro-rated Bank Payment Channel Promotion  for return refund Items': 'pro_rated_bank_payment_channel_promotion_for_return_refund_item',   # alias double spasi (2023)
    'Pro-rated Shopee Payment Channel Promotion for return refund Items': 'pro_rated_shopee_payment_channel_promotion_for_return_refund_it',
    'Pro-rated Shopee Payment Channel Promotion  for return refund Items': 'pro_rated_shopee_payment_channel_promotion_for_return_refund_it'  # alias double spasi (2023)
}

SHOPEE_INCOME_SF_COLUMN_MAP = {
    'No.': 'no',
    'No. Pesanan': 'no_pesanan',
    'Biaya Layanan Cashback XTRA': 'biaya_layanan_cashback_xtra',
    'Biaya Layanan CashbackXTRA': 'biaya_layanan_cashbackxtra',     # kolom berbeda (tanpa spasi, 2024)
    'Biaya Layanan Cashback Xtra': 'biaya_layanan_cashback_xtra',   # alias kapitalisasi berbeda (2023 official)
    'Biaya Layanan Gratis Ongkir XTRA': 'biaya_layanan_gratis_ongkir_xtra',
    'Biaya Layanan Gratis Ongkir Xtra': 'biaya_layanan_gratis_ongkir_xtra_2',
    'Biaya Layanan Promo XTRA': 'biaya_layanan_promo_xtra',
    'Biaya Pembayaran': 'biaya_pembayaran',
    'Biaya Program Shopee Live Xtra': 'biaya_program_shopee_live_xtra',
    'Biaya Campaign 1.1 2024': 'biaya_campaign_1_1',
    'Biaya Campaign 2.2 2024': 'biaya_campaign_2_2',
    'Biaya Campaign 3.3 2024': 'biaya_campaign_3_3',
    'Biaya Campaign 4.4 2024': 'biaya_campaign_4_4',
    'Biaya Campaign 5.5 2024': 'biaya_campaign_5_5',
    'Biaya Campaign 6.6 2024': 'biaya_campaign_6_6',
    'Biaya Campaign 7.7 2024': 'biaya_campaign_7_7',
    'Biaya Campaign 8.8 2024': 'biaya_campaign_8_8',
    'Biaya Campaign 9.9 2024': 'biaya_campaign_9_9',
    'Biaya Campaign 10.10 2024': 'biaya_campaign_10_10',
    'Biaya Campaign 11.11 2024': 'biaya_campaign_11_11',
    'Biaya Campaign 12.12 2024': 'biaya_campaign_12_12',
    'Biaya Campaign 1.1 2025': 'biaya_campaign_1_1',
    'Biaya Campaign 2.2 2025': 'biaya_campaign_2_2',
    'Biaya Campaign 3.3 2025': 'biaya_campaign_3_3',
    'Biaya Campaign 4.4 2025': 'biaya_campaign_4_4',
    'Biaya Campaign 5.5 2025': 'biaya_campaign_5_5',
    'Biaya Campaign 6.6 2025': 'biaya_campaign_6_6',
    'Biaya Campaign 7.7 2025': 'biaya_campaign_7_7',
    'Biaya Campaign 8.8 2025': 'biaya_campaign_8_8',
    'Biaya Campaign 9.9 2025': 'biaya_campaign_9_9',
    'Biaya Campaign 10.10 2025': 'biaya_campaign_10_10',
    'Biaya Campaign 11.11 2025': 'biaya_campaign_11_11',
    'Biaya Campaign 12.12 2025': 'biaya_campaign_12_12',
}

SHOPEE_INCOME_SD_COLUMN_MAP = {
    'No. Pesanan': 'no_pesanan',
    'Estimasi Ongkos Kirim:': 'estimasi_ongkos_kirim', 
    'Ongkos Kirim yang Dibayarkan ke Jasa Kirim:': 'ongkos_kirim_yang_dibayarkan_ke_jasa_kirim', 
    'Discrepancy reason': 'discrepancy_reason'
}

SHOPEE_INCOME_ADJ_COLUMN_MAP = {
    'No.': 'no',
    'Tanggal Penyesuaian Dibuat': 'tanggal_penyesuaian_dibuat',
    'Tipe Penyesuaian | Deskripsi': 'tipe_penyesuaian_deskripsi',
    'Alasan Penyesuaian': 'alasan_penyesuaian',
    'Biaya Penyesuaian': 'biaya_penyesuaian',
    'No. Pesanan Terhubung': 'no_pesanan_terhubung',
    'Tanggal Dana Dilepaskan': 'tanggal_dana_dilepaskan'
}

LAZADA_INCOME_COLUMN_MAP = {
    'Periode Laporan': 'periode_laporan', 
    'Nomor Laporan': 'nomor_laporan', 
    'Tanggal Transaksi': 'tanggal_transaksi', 
    'Nama Biaya': 'nama_biaya', 
    'Jumlah (Termasuk Pajak)': 'jumlah_termasuk_pajak', 
    'VAT Amount': 'vat_amount', 
    'Status Pelepasan Dana': 'status_pelepasan_dana', 
    'Tanggal Dilepas': 'tanggal_dilepas', 
    'Komentar': 'komentar', 
    'Tanggal Pesanan Dibuat': 'tanggal_pesanan_dibuat', 
    'Nomor Pesanan': 'nomor_pesanan', 
    'ID Pesanan': 'id_pesanan', 
    'SKU Penjual': 'sku_penjual', 
    'Lazada SKU': 'lazada_sku', 
    'WHT Amount': 'wht_amount', 
    'WHT termasuk dalam jumlah': 'wht_termasuk_dalam_jumlah', 
    'Status Pesanan': 'status_pesanan', 
    'Nama Produk': 'nama_produk', 
    'Short Code': 'short_code'
}


TIKTOK_REPORT_COLUMN_MAP = {
    'Type': 'type',
    'Reference ID': 'reference_id',
    'Request time': 'request_time',
    'Amount': 'amount',
    'Status': 'status',
    'Success time': 'success_time',
    'Bank account': 'bank_account'
}

SHOPEE_REPORT_COLUMN_MAP = {
    'Tanggal Transaksi': 'tanggal_transaksi',
    'Tipe Transaksi': 'tipe_transaksi',
    'Deskripsi': 'deskripsi',
    'No. Pesanan': 'no_pesanan',
    'Jenis Transaksi': 'jenis_transaksi',
    'Jumlah': 'jumlah',
    'Status': 'status',
    'Saldo Akhir': 'saldo_akhir'
}

LAZADA_REPORT_COLUMN_MAP = {
    'Transaction Number': 'transaction_number',
    'Transaction Time': 'transaction_time',
    'Type': 'type',
    'Sub Type': 'sub_type',
    'Amount': 'amount',
    'Remarks': 'remarks'
}

CREWDIBLE_COLUMN_MAP = {
    'No': 'no',
    'Gudang': 'gudang',
    'Tanggal Transaksi': 'tanggal_transaksi',
    'No. Transaksi': 'no_transaksi',
    'Status': 'status',
    'Pengirim': 'pengirim',
    'No. HP Pengirim': 'no_hp_pengirim',
    'Nama Toko': 'nama_toko',
    'Nama Marketplace': 'nama_marketplace',
    'Penerima': 'penerima',
    'No. HP Penerima': 'no_hp_penerima',
    'Alamat Penerima': 'alamat_penerima',
    'Nama Produk': 'nama_produk',
    'No. SKU': 'no_sku',
    'Qty Produk': 'qty_produk',
    'Harga Produk': 'harga_produk',
    'Total Harga Produk': 'total_harga_produk',
    'Total Nilai Transaksi': 'total_nilai_transaksi',
    'Biaya Transaksi': 'biaya_transaksi',
    'PPN Biaya Transaksi': 'ppn_biaya_transaksi',
    'Material Packaging': 'material_packaging',
    'Harga Material Packaging': 'harga_material_packaging',
    'Qty Material Packaging': 'qty_material_packaging',
    'Total Harga Material Packaging': 'total_harga_material_packaging',
    'Total Biaya Packaging': 'total_biaya_packaging',
    'PPN Total Biaya Packaging': 'ppn_total_biaya_packaging',
    'Total Biaya QC': 'total_biaya_qc',
    'PPN Total Biaya QC': 'ppn_total_biaya_qc',
    'Total Biaya Shipping Label': 'total_biaya_shipping_label',
    'PPN Total Biaya Shipping Label': 'ppn_total_biaya_shipping_label',
    'Logistik': 'logistik',
    'Kode Booking': 'kode_booking',
    'No. AWB': 'no_awb',
    'Biaya Logistik': 'biaya_logistik',
    'Total Biaya Transaksi': 'total_biaya_transaksi',
}

VALID_CREWDIBLE_COLS = [
    'no', 'gudang', 'tanggal_transaksi', 'no_transaksi', 'status',
    'pengirim', 'no_hp_pengirim', 'nama_toko', 'nama_marketplace',
    'penerima', 'no_hp_penerima', 'alamat_penerima', 'nama_produk',
    'no_sku', 'qty_produk', 'harga_produk', 'total_harga_produk',
    'total_nilai_transaksi', 'biaya_transaksi', 'ppn_biaya_transaksi',
    'material_packaging', 'harga_material_packaging', 'qty_material_packaging',
    'total_harga_material_packaging', 'total_biaya_packaging', 'ppn_total_biaya_packaging',
    'total_biaya_qc', 'ppn_total_biaya_qc', 'total_biaya_shipping_label',
    'ppn_total_biaya_shipping_label', 'logistik', 'kode_booking', 'no_awb',
    'biaya_logistik', 'total_biaya_transaksi', 'source_filename',
]


# ==========================================
# FUNGSI PEMROSES FASE: ORDER
# ==========================================
def process_order_file(file_path, marketplace, engine, nama_toko_override=None):
    """
    Membaca file ORDER, mencuci datanya, dan mengirimkannya ke database.
    """
    filename = os.path.basename(file_path)
    logger.info(f"Mengekstrak file ORDER: {filename} [{marketplace.upper()}]")
    
    # 1. Baca File Mentah (Anggap semua sebagai Text/String agar tidak crash)
    try:
        if file_path.endswith('.csv'):
            # Beberapa CSV e-commerce menggunakan pemisah titik koma (;), tapi pandas otomatis koma. 
            # Jika nanti ada error baca CSV, kita bisa tambahkan parameter sep=';' atau error_bad_lines
            df = pd.read_csv(file_path, dtype=str) 
        elif marketplace == 'tiktok_tokopedia':
            # pd.read_excel() tidak bisa membaca file TikTok/Tokopedia dengan benar,
            # sehingga dibaca langsung via openpyxl lalu dikonversi ke DataFrame.
            # Baris ke-2 (index 1) adalah deskripsi kolom, dilewati dengan slice [2:]
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            rows = list(ws.values)
            cols = rows[0]
            df = pd.DataFrame(rows[2:], columns=cols).astype(str)
            df = df.replace('None', pd.NA)
        else:
            df = pd.read_excel(file_path, dtype=str, engine='openpyxl')
    except Exception as e:
        logger.error(f"Gagal membaca file {filename}: {e}")
        return

    # 2. Suntik Metadata
    df['nama_toko'] = nama_toko_override if nama_toko_override else extract_store_name(filename)
    df['source_filename'] = filename

    # 3. Bersihkan spasi kosong di pinggiran nama kolom Excel (sangat penting untuk Dictionary)
    df.columns = df.columns.str.strip()

    # 4. Ganti Nama Kolom dengan Dictionary & Filter Skema
    if marketplace == 'shopee':
        # UNIQUE_COLS = ['no_pesanan', 'nomor_referensi_sku', 'nama_variasi']
        df = df.rename(columns=SHOPEE_ORDER_COLUMN_MAP)
        df = enforce_schema(df, VALID_SHOPEE_ORDER_COLS, filename, 'stg_shopee_orders', file_path, engine)
        # df = dedup_df(df, UNIQUE_COLS, 'stg_shopee_orders', filename)
        # df = filter_existing_rows(df, UNIQUE_COLS, 'stg_shopee_orders', engine, filename)
        # if df.empty:
        #     return
        load_to_staging(df, 'stg_shopee_orders', engine)

    elif marketplace == 'lazada':
        # UNIQUE_COLS = ['order_item_id']
        df = df.rename(columns=LAZADA_ORDER_COLUMN_MAP)
        df = enforce_schema(df, VALID_LAZADA_ORDER_COLS, filename, 'stg_lazada_orders', file_path, engine)
        # df = dedup_df(df, UNIQUE_COLS, 'stg_lazada_orders', filename)
        # df = filter_existing_rows(df, UNIQUE_COLS, 'stg_lazada_orders', engine, filename)
        # if df.empty:
        #     return
        load_to_staging(df, 'stg_lazada_orders', engine)

    elif marketplace == 'tiktok_tokopedia':
        # UNIQUE_COLS = ['order_id', 'seller_sku', 'product_name']
        df = df.rename(columns=TIKTOK_ORDER_COLUMN_MAP)
        df = enforce_schema(df, VALID_TIKTOK_ORDER_COLS, filename, 'stg_tiktok_tokopedia_orders', file_path, engine)
        # df = dedup_df(df, UNIQUE_COLS, 'stg_tiktok_tokopedia_orders', filename)
        # df = filter_existing_rows(df, UNIQUE_COLS, 'stg_tiktok_tokopedia_orders', engine, filename)
        # if df.empty:
        #     return
        load_to_staging(df, 'stg_tiktok_tokopedia_orders', engine)


# ==========================================
# FUNGSI PEMROSES FASE: INCOME
# ==========================================
def process_income_file(file_path, marketplace, engine, nama_toko_override=None):
    """
    Membaca file INCOME, menangani multi-sheet (Shopee) & format dinamis,
    mencuci datanya, dan mengirimkannya ke database.
    """
    filename = os.path.basename(file_path)
    logger.info(f"Mengekstrak file INCOME: {filename} [{marketplace.upper()}]")

    nama_toko = nama_toko_override if nama_toko_override else extract_store_name(filename)

    try:
        if marketplace in ['tiktok', 'tiktok_tokopedia']:
            # Logika khusus TikTok: Cek apakah ada sheet 'Order details'
            xl = pd.ExcelFile(file_path, engine='openpyxl')
            target_sheet = 'Order details' if 'Order details' in xl.sheet_names else 0
            
            df = pd.read_excel(file_path, sheet_name=target_sheet, dtype=str, engine='openpyxl')
            df['nama_toko'] = nama_toko
            df['source_filename'] = filename
            df.columns = df.columns.astype(str).str.strip()
            
            df = df.rename(columns=TIKTOK_INCOME_COLUMN_MAP)
            df = enforce_schema(df, VALID_TIKTOK_INCOME_COLS, filename, 'stg_tiktok_tokopedia_income', file_path, engine)
            load_to_staging(df, 'stg_tiktok_tokopedia_income', engine)

        elif marketplace == 'lazada':
            # Lazada format standar (header di baris 1)
            df = pd.read_excel(file_path, dtype=str, engine='openpyxl')
            df['nama_toko'] = nama_toko
            df['source_filename'] = filename
            df.columns = df.columns.astype(str).str.strip()
            
            df = df.rename(columns=LAZADA_INCOME_COLUMN_MAP)
            df = enforce_schema(df, VALID_LAZADA_INCOME_COLS, filename, 'stg_lazada_income', file_path, engine)
            load_to_staging(df, 'stg_lazada_income', engine)

        elif marketplace == 'shopee':
            # Shopee Multi-Sheet: gunakan openpyxl langsung agar semua kolom terbaca dengan benar
            wb = openpyxl.load_workbook(file_path, data_only=True)

            for sheet in wb.sheetnames:
                sheet_lower = sheet.lower()
                ws = wb[sheet]
                rows = list(ws.values)

                if not rows:
                    continue

                # 1a. SELLER FEE (format 2023: gabungan OPF + Service Fee, header di row 2, index 1)
                if 'seller fee' in sheet_lower:
                    if len(rows) < 3:
                        continue
                    cols = [str(c).strip() if c is not None else None for c in rows[1]]
                    df = pd.DataFrame(rows[2:], columns=cols).astype(str)
                    df = df.replace('None', pd.NA)

                    # Kolom B tidak punya nama di 2023 → rename by position ke 'Lihat berdasarkan'
                    df_cols = list(df.columns)
                    if len(df_cols) > 1 and df_cols[1] is None:
                        df_cols[1] = 'Lihat berdasarkan'
                        df.columns = df_cols

                    df['nama_toko'] = nama_toko
                    df['source_filename'] = filename
                    df.columns = df.columns.astype(str).str.strip()

                    # Load ke OPF (enforce_schema buang kolom biaya layanan)
                    df_opf = df.rename(columns=SHOPEE_INCOME_OPF_COLUMN_MAP)
                    df_opf = enforce_schema(df_opf, VALID_SHOPEE_INCOME_OPF_COLS, f"{filename} ({sheet})", 'stg_shopee_income_order_processing_fee', file_path, engine)
                    load_to_staging(df_opf, 'stg_shopee_income_order_processing_fee', engine)

                    # Load ke Service Fee (enforce_schema buang kolom OPF)
                    df_sf = df.rename(columns=SHOPEE_INCOME_SF_COLUMN_MAP)
                    df_sf = enforce_schema(df_sf, VALID_SHOPEE_INCOME_SF_COLS, f"{filename} ({sheet})", 'stg_shopee_income_service_fee', file_path, engine)
                    load_to_staging(df_sf, 'stg_shopee_income_service_fee', engine)

                # 1b. ORDER PROCESSING FEE (format 2025: header di row 1, index 0)
                elif 'order processing fee' in sheet_lower:
                    cols = [str(c).strip() if c is not None else c for c in rows[0]]
                    df = pd.DataFrame(rows[1:], columns=cols).astype(str)
                    df = df.replace('None', pd.NA)
                    df['nama_toko'] = nama_toko
                    df['source_filename'] = filename
                    df.columns = df.columns.astype(str).str.strip()
                    df = df.rename(columns=SHOPEE_INCOME_OPF_COLUMN_MAP)
                    df = enforce_schema(df, VALID_SHOPEE_INCOME_OPF_COLS, f"{filename} ({sheet})", 'stg_shopee_income_order_processing_fee', file_path, engine)
                    load_to_staging(df, 'stg_shopee_income_order_processing_fee', engine)

                # 2. INCOME MAIN (5 baris metadata, header di row 6, index 5)
                elif 'income' in sheet_lower:
                    if len(rows) < 7:
                        continue
                    cols = [str(c).strip() if c is not None else c for c in rows[5]]
                    df = pd.DataFrame(rows[6:], columns=cols).astype(str)
                    df = df.replace('None', pd.NA)
                    if df.empty or len(df.columns) < 5:
                        continue
                    df['nama_toko'] = nama_toko
                    df['source_filename'] = filename
                    df.columns = df.columns.astype(str).str.strip()
                    df = df.rename(columns=SHOPEE_INCOME_MAIN_COLUMN_MAP)
                    df = enforce_schema(df, VALID_SHOPEE_INCOME_MAIN_COLS, f"{filename} ({sheet})", 'stg_shopee_income_main', file_path, engine)
                    load_to_staging(df, 'stg_shopee_income_main', engine)

                # 3. SERVICE FEE DETAILS (merged header di row 1, header di row 2, index 1)
                elif 'service fee' in sheet_lower:
                    if len(rows) < 3:
                        continue
                    cols = [str(c).strip() if c is not None else c for c in rows[1]]
                    df = pd.DataFrame(rows[2:], columns=cols).astype(str)
                    df = df.replace('None', pd.NA)
                    df['nama_toko'] = nama_toko
                    df['source_filename'] = filename
                    df.columns = df.columns.astype(str).str.strip()
                    df = df.rename(columns=SHOPEE_INCOME_SF_COLUMN_MAP)
                    df = enforce_schema(df, VALID_SHOPEE_INCOME_SF_COLS, f"{filename} ({sheet})", 'stg_shopee_income_service_fee', file_path, engine)
                    load_to_staging(df, 'stg_shopee_income_service_fee', engine)

                # 4. SHIPPING FEE DISCREPANCY (catatan di row 1, header di row 2, index 1)
                elif 'shipping fee' in sheet_lower:
                    if len(rows) < 3:
                        continue
                    cols = [str(c).strip() if c is not None else c for c in rows[1]]
                    df = pd.DataFrame(rows[2:], columns=cols).astype(str)
                    df = df.replace('None', pd.NA)
                    df['nama_toko'] = nama_toko
                    df['source_filename'] = filename
                    df.columns = df.columns.astype(str).str.strip()
                    df = df.rename(columns=SHOPEE_INCOME_SD_COLUMN_MAP)
                    df = enforce_schema(df, VALID_SHOPEE_INCOME_SD_COLS, f"{filename} ({sheet})", 'stg_shopee_income_shipping_discrepancy', file_path, engine)
                    load_to_staging(df, 'stg_shopee_income_shipping_discrepancy', engine)

                # 5. ADJUSTMENT (DYNAMIC ANCHOR)
                elif 'adjustment' in sheet_lower:
                    header_idx = -1

                    # Cari baris yang mengandung 'No.' dan 'Tanggal' (atau 'Tipe Penyesuaian')
                    for i, row in enumerate(rows):
                        col_0 = str(row[0]).strip() if row[0] is not None else ''
                        col_1_to_3 = " ".join([str(x) for x in row[1:4]]).lower()
                        if col_0 == 'No.' and ('tanggal' in col_1_to_3 or 'tipe' in col_1_to_3):
                            header_idx = i
                            break

                    if header_idx != -1:
                        cols = [str(c).strip() if c is not None else c for c in rows[header_idx]]
                        df = pd.DataFrame(rows[header_idx + 1:], columns=cols).astype(str)
                        df = df.replace('None', pd.NA)

                        # Bersihkan kolom None yang mungkin terbawa
                        df = df.loc[:, df.columns.notna()]

                        df['nama_toko'] = nama_toko
                        df['source_filename'] = filename
                        df.columns = df.columns.astype(str).str.strip()
                        df = df.rename(columns=SHOPEE_INCOME_ADJ_COLUMN_MAP)

                        # ==========================================
                        # VAKSIN ANTI "TOTAL BIAYA" & BARIS KOSONG
                        # ==========================================
                        # 1. Hapus baris jika kolom 'no' kosong (NaN/None) akibat sisa formatting Excel
                        df = df.dropna(subset=['no'])
                        # 2. Hapus baris jika kolom 'no' berisi kata "Total" (mengabaikan huruf besar/kecil)
                        df = df[~df['no'].astype(str).str.lower().str.contains('total', na=False)]
                        # ==========================================

                        df = enforce_schema(df, VALID_SHOPEE_INCOME_ADJ_COLS, f"{filename} ({sheet})", 'stg_shopee_income_adjustment', file_path, engine)
                        load_to_staging(df, 'stg_shopee_income_adjustment', engine)
                    else:
                        logger.warning(f"⚠️ Header 'No.' tidak ditemukan di sheet {sheet} pada file {filename}. Sheet diabaikan.")

                # Sheet tidak cocok dengan pola manapun → kemungkinan tipe baru dari Shopee
                else:
                    logger.warning(f"⚠️ Sheet tidak dikenal diabaikan: '{sheet}' pada file {filename}")
                    notify_unknown_sheet(filename, sheet, file_path, engine)

    except Exception as e:
        logger.error(f"❌ Gagal memproses file INCOME {filename}: {e}")


# ==========================================
# FUNGSI PEMROSES FASE: REPORT
# ==========================================
def process_report_file(file_path, marketplace, engine, nama_toko_override=None):
    """
    Membaca file REPORT dan mengirimkannya ke database.
    - TikTok: Sheet1, header di row 0, data mulai row 1
    - Lazada: Sheet 'Balance Transactions', header di row 0, data mulai row 1
    """
    filename = os.path.basename(file_path)
    logger.info(f"Mengekstrak file REPORT: {filename} [{marketplace.upper()}]")

    if nama_toko_override:
        nama_toko = nama_toko_override
    elif marketplace == 'tiktok_tokopedia':
        nama_toko = extract_tiktok_report_store_name(filename)
    else:
        nama_toko = extract_store_name(filename)

    try:
        if marketplace == 'tiktok_tokopedia':
            if file_path.endswith('.csv'):
                df = pd.read_csv(file_path, dtype=str).fillna('')
                df.columns = df.columns.astype(str).str.strip()
            else:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                ws = wb.active
                rows = list(ws.values)
                if not rows:
                    logger.warning(f"⚠️ File kosong: {filename}")
                    return
                cols = [str(c).strip() if c is not None else None for c in rows[0]]
                df = pd.DataFrame(rows[1:], columns=cols).astype(str)
                df = df.replace('None', pd.NA)

            df['nama_toko'] = nama_toko
            df['source_filename'] = filename
            df.columns = df.columns.astype(str).str.strip()

            df = df.rename(columns=TIKTOK_REPORT_COLUMN_MAP)
            df = enforce_schema(df, VALID_TIKTOK_REPORT_COLS, filename, 'stg_tiktok_tokopedia_report', file_path, engine)
            load_to_staging(df, 'stg_tiktok_tokopedia_report', engine)

        elif marketplace == 'lazada':
            wb = openpyxl.load_workbook(file_path, data_only=True)
            if 'Balance Transactions' not in wb.sheetnames:
                logger.warning(f"⚠️ Sheet 'Balance Transactions' tidak ditemukan di {filename}. File diabaikan.")
                return

            ws = wb['Balance Transactions']
            rows = list(ws.values)
            if not rows:
                logger.warning(f"⚠️ File kosong: {filename}")
                return

            cols = [str(c).strip() if c is not None else None for c in rows[0]]
            df = pd.DataFrame(rows[1:], columns=cols).astype(str)
            df = df.replace('None', pd.NA)

            df['nama_toko'] = nama_toko
            df['source_filename'] = filename
            df.columns = df.columns.astype(str).str.strip()

            df = df.rename(columns=LAZADA_REPORT_COLUMN_MAP)
            df = enforce_schema(df, VALID_LAZADA_REPORT_COLS, filename, 'stg_lazada_report', file_path, engine)
            load_to_staging(df, 'stg_lazada_report', engine)

        elif marketplace == 'shopee':
            wb = openpyxl.load_workbook(file_path, data_only=True)
            if 'Transaction Report' not in wb.sheetnames:
                logger.warning(f"⚠️ Sheet 'Transaction Report' tidak ditemukan di {filename}. File diabaikan.")
                return

            ws = wb['Transaction Report']
            rows = list(ws.values)

            # Header ada di row index 17, data mulai row index 18
            # (17 baris pertama adalah metadata: username, tanggal, ringkasan saldo)
            if len(rows) < 19:
                logger.warning(f"⚠️ File terlalu pendek atau kosong: {filename}")
                return

            cols = [str(c).strip() if c is not None else None for c in rows[17]]
            df = pd.DataFrame(rows[18:], columns=cols).astype(str)
            df = df.replace('None', pd.NA)

            df['nama_toko'] = nama_toko
            df['source_filename'] = filename
            df.columns = df.columns.astype(str).str.strip()

            df = df.rename(columns=SHOPEE_REPORT_COLUMN_MAP)
            df = enforce_schema(df, VALID_SHOPEE_REPORT_COLS, filename, 'stg_shopee_report', file_path, engine)
            load_to_staging(df, 'stg_shopee_report', engine)

        else:
            logger.warning(f"⚠️ Marketplace '{marketplace}' tidak memiliki handler REPORT. File diabaikan: {filename}")

    except Exception as e:
        logger.error(f"❌ Gagal memproses file REPORT {filename}: {e}")


# ==========================================
# FUNGSI PEMROSES FASE: CREWDIBLE
# ==========================================
_XLSX_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'

def _read_crewdible_xml_patched(file_path):
    """
    Fallback reader untuk file .xls dari Crewdible yang sebenarnya adalah xlsx
    namun mengandung XML invalid (nilai '-' pada sel numerik dan tidak ada
    sharedStrings — semua string disimpan sebagai inlineStr).

    Strategi:
    1. Buka sebagai ZIP
    2. Patch XML: ganti <v>-</v> dengan <v>0</v>
    3. Parse sel tipe 'inlineStr' via <is><t>...</t></is>
    4. Parse sel numerik via <v>...</v>
    """
    with zipfile.ZipFile(file_path, 'r') as z:
        xml_raw = z.read('xl/worksheets/sheet1.xml').decode('utf-8')

    xml_raw = xml_raw.replace('<v>-</v>', '<v>0</v>')
    root = ET.fromstring(xml_raw)

    rows_data = []
    for row_el in root.iter(f'{{{_XLSX_NS}}}row'):
        row = []
        for cell in row_el.iter(f'{{{_XLSX_NS}}}c'):
            t = cell.get('t', 'n')
            if t == 'inlineStr':
                is_el = cell.find(f'{{{_XLSX_NS}}}is')
                t_el = is_el.find(f'{{{_XLSX_NS}}}t') if is_el is not None else None
                val = t_el.text if t_el is not None else None
            else:
                v_el = cell.find(f'{{{_XLSX_NS}}}v')
                val = v_el.text if v_el is not None else None
            row.append(val)
        rows_data.append(row)

    return rows_data


def process_crewdible_file(file_path, engine):
    """
    Membaca file transaksi Crewdible (3PL fulfillment), mencuci datanya,
    dan mengirimkannya ke staging.stg_crewdible.

    Mendukung dua format:
    - .xlsx standar          → dibaca langsung via openpyxl
    - .xls (xlsx tersamar)   → fallback ke _read_crewdible_xml_patched()
      (file Crewdible 2024 bulanan & 2026 menggunakan format ini:
       inlineStr tanpa sharedStrings, nilai '-' pada sel numerik)
    """
    filename = os.path.basename(file_path)
    logger.info(f"Mengekstrak file CREWDIBLE: {filename}")

    try:
        # Coba openpyxl dulu (bekerja untuk .xlsx standar)
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        rows = list(ws.values)
        if not rows:
            logger.warning(f"⚠️ File kosong: {filename}")
            return

        # Auto-detect header row: cari baris yang mengandung 'No' di kolom pertama
        header_idx = 0
        for i, row in enumerate(rows[:5]):
            if row and str(row[0]).strip() == 'No':
                header_idx = i
                break

        cols = [str(c).strip() if c is not None else f'_col{i}' for i, c in enumerate(rows[header_idx])]
        df = pd.DataFrame(rows[header_idx + 1:], columns=cols).astype(str)
        df = df.replace('None', pd.NA)

    except Exception:
        # Fallback: file .xls dengan XML invalid (format Crewdible 2024 bulanan / 2026)
        logger.warning(f"⚠️ openpyxl gagal untuk {filename}, mencoba XML patching...")
        try:
            rows_data = _read_crewdible_xml_patched(file_path)
            if not rows_data:
                logger.warning(f"⚠️ File kosong setelah XML patching: {filename}")
                return

            header_idx = 0
            for i, row in enumerate(rows_data[:5]):
                if row and str(row[0]).strip() == 'No':
                    header_idx = i
                    break

            cols = [str(c).strip() if c else f'_col{i}' for i, c in enumerate(rows_data[header_idx])]
            df = pd.DataFrame(rows_data[header_idx + 1:], columns=cols)
            df = df.replace([None, ''], pd.NA)
            df = df.astype(str).replace('nan', pd.NA)

        except Exception as e2:
            logger.error(f"❌ Gagal membaca file {filename} (openpyxl + XML patching): {e2}")
            return

    # =====================================================================
    # FORWARD-FILL: Satu transaksi bisa punya >1 baris jika ada >1 jenis
    # Material Packaging. Baris ke-2 dst. hanya mengisi kolom packaging;
    # semua kolom level-transaksi lainnya (No. Transaksi, Nama Toko, dll.)
    # diisi dari baris pertama transaksi tersebut agar setiap baris menjadi
    # self-contained di staging.
    # =====================================================================
    FFILL_COLS = [
        'No', 'Gudang', 'Tanggal Transaksi', 'No. Transaksi', 'Status',
        'Pengirim', 'No. HP Pengirim', 'Nama Toko', 'Nama Marketplace',
        'Penerima', 'No. HP Penerima', 'Alamat Penerima',
        'Nama Produk', 'No. SKU', 'Qty Produk', 'Harga Produk',
        'Total Harga Produk', 'Total Nilai Transaksi',
        'Biaya Transaksi', 'PPN Biaya Transaksi',
        'Total Biaya Packaging', 'PPN Total Biaya Packaging',
        'Total Biaya QC', 'PPN Total Biaya QC',
        'Total Biaya Shipping Label', 'PPN Total Biaya Shipping Label',
        'Logistik', 'Kode Booking', 'No. AWB', 'Biaya Logistik',
        'Total Biaya Transaksi',
    ]
    cols_to_fill = [c for c in FFILL_COLS if c in df.columns]
    df[cols_to_fill] = df[cols_to_fill].ffill()

    # Setelah forward-fill, buang baris yang tidak memiliki material packaging
    # (baris summary/total/kosong di akhir file atau antar transaksi)
    df = df.dropna(subset=['Material Packaging'])
    if df.empty:
        logger.warning(f"⚠️ File {filename} tidak mengandung data valid setelah filter. Dilewati.")
        return

    # Metadata
    df['source_filename'] = filename

    # Bersihkan spasi di nama kolom
    df.columns = df.columns.str.strip()

    # Rename kolom ke snake_case
    df = df.rename(columns=CREWDIBLE_COLUMN_MAP)

    # Schema guard
    df = enforce_schema(df, VALID_CREWDIBLE_COLS, filename, 'stg_crewdible', file_path, engine)

    # Dedup dalam file: kunci = no_transaksi + no_sku + material_packaging + qty_material_packaging
    # Qty dimasukkan ke kunci agar dua batch packing yang berbeda jumlah tidak ikut dibuang.
    # Hanya baris yang benar-benar 100% identik pada keempat kolom ini yang dihapus.
    # UNIQUE_COLS = ['no_transaksi', 'no_sku', 'material_packaging', 'qty_material_packaging']
    # df = dedup_df(df, UNIQUE_COLS, 'stg_crewdible', filename)

    # Filter lintas-file: cek kunci unik di seluruh tabel,
    # bukan hanya untuk source_filename ini — mencegah duplikat antara file bulanan.
    UNIQUE_COLS = ['no_transaksi', 'no_sku', 'material_packaging', 'qty_material_packaging']
    df = filter_existing_rows(df, UNIQUE_COLS, 'stg_crewdible', engine, filename, global_check=True)
    if df.empty:
        return

    load_to_staging(df, 'stg_crewdible', engine)