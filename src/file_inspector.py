"""
src/file_inspector.py

Bertugas menghitung jumlah baris yang akan dimuat ke tabel staging primer
dari sebuah file spreadsheet, tanpa benar-benar memuatnya ke database.

Logika pembacaan file di sini mereplikasi persis logika di extract_loader.py
agar angka yang dihasilkan selalu konsisten dengan yang akan masuk ke DB.
"""

import openpyxl
import pandas as pd
from sqlalchemy import text
from src.db_config import logger


# ── Konfigurasi order ID per (fase, marketplace) ───────────────────────────────
# (excel_col, staging_col, staging_table)
ORDER_ID_CONFIG = {
    ('ORDER',  'shopee'):           ('No. Pesanan',        'no_pesanan',           'stg_shopee_orders'),
    ('ORDER',  'tiktok_tokopedia'): ('Order ID',           'order_id',             'stg_tiktok_tokopedia_orders'),
    ('ORDER',  'lazada'):           ('orderNumber',        'order_number',         'stg_lazada_orders'),
    ('INCOME', 'shopee'):           ('No. Pesanan',        'no_pesanan',           'stg_shopee_income_main'),
    ('INCOME', 'tiktok_tokopedia'): ('Order/adjustment ID','order_adjustment_id',  'stg_tiktok_tokopedia_income'),
    ('INCOME', 'lazada'):           ('Nomor Pesanan',      'nomor_pesanan',        'stg_lazada_income'),
    ('REPORT', 'shopee'):           ('No. Pesanan',        'no_pesanan',           'stg_shopee_report'),
    ('REPORT', 'tiktok_tokopedia'): ('Reference ID',       'reference_id',         'stg_tiktok_tokopedia_report'),
    ('REPORT', 'lazada'):           ('Transaction Number', 'transaction_number',   'stg_lazada_report'),
}

# ============================================================
# MAPPING: (fase, marketplace) → tabel staging primer
# ============================================================
PRIMARY_TABLE = {
    ('ORDER',  'shopee'):           'stg_shopee_orders',
    ('ORDER',  'tiktok_tokopedia'): 'stg_tiktok_tokopedia_orders',
    ('ORDER',  'lazada'):           'stg_lazada_orders',
    ('INCOME', 'shopee'):           'stg_shopee_income_main',
    ('INCOME', 'tiktok_tokopedia'): 'stg_tiktok_tokopedia_income',
    ('INCOME', 'lazada'):           'stg_lazada_income',
    ('REPORT', 'shopee'):           'stg_shopee_report',
    ('REPORT', 'tiktok_tokopedia'): 'stg_tiktok_tokopedia_report',
    ('REPORT', 'lazada'):           'stg_lazada_report',
}


# ============================================================
# FUNGSI PENGHITUNG BARIS PER (FASE, MARKETPLACE)
# ============================================================

def _count_shopee_order(file_path):
    df = pd.read_excel(file_path, dtype=str, engine='openpyxl')
    return len(df)


def _count_tiktok_order(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.values)
    wb.close()
    if len(rows) < 3:
        return 0
    # Baris index 0 = header, index 1 = deskripsi kolom (dilewati), data mulai index 2
    data_rows = [r for r in rows[2:] if any(v is not None for v in r)]
    return len(data_rows)


def _count_lazada_order(file_path):
    df = pd.read_excel(file_path, dtype=str, engine='openpyxl')
    return len(df)


def _count_shopee_income(file_path):
    """
    Shopee INCOME bisa punya banyak sheet Income (Income, Income - 1, Income - 2, dst.).
    Semua sheet yang mengandung kata 'income' dijumlahkan — sama persis dengan
    logika process_income_file() di extract_loader.py.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    total = 0
    for sheet in wb.sheetnames:
        if 'income' not in sheet.lower():
            continue
        ws = wb[sheet]
        rows = list(ws.values)
        if len(rows) < 7:
            continue
        # Header di row index 5, data mulai index 6
        data_rows = [r for r in rows[6:] if any(v is not None for v in r)]
        total += len(data_rows)
    wb.close()
    return total


def _count_tiktok_income(file_path):
    xl = pd.ExcelFile(file_path, engine='openpyxl')
    target_sheet = 'Order details' if 'Order details' in xl.sheet_names else 0
    df = pd.read_excel(file_path, sheet_name=target_sheet, dtype=str, engine='openpyxl')
    return len(df)


def _count_lazada_income(file_path):
    df = pd.read_excel(file_path, dtype=str, engine='openpyxl')
    return len(df)


def _count_shopee_report(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    if 'Transaction Report' not in wb.sheetnames:
        wb.close()
        return 0
    ws = wb['Transaction Report']
    rows = list(ws.values)
    wb.close()
    if len(rows) < 19:
        return 0
    # Header di row index 17, data mulai index 18
    data_rows = [r for r in rows[18:] if any(v is not None for v in r)]
    return len(data_rows)


def _count_tiktok_report(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.values)
    wb.close()
    if len(rows) < 2:
        return 0
    # Header di row index 0, data mulai index 1
    data_rows = [r for r in rows[1:] if any(v is not None for v in r)]
    return len(data_rows)


def _count_lazada_report(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    if 'Balance Transactions' not in wb.sheetnames:
        wb.close()
        return 0
    ws = wb['Balance Transactions']
    rows = list(ws.values)
    wb.close()
    if len(rows) < 2:
        return 0
    data_rows = [r for r in rows[1:] if any(v is not None for v in r)]
    return len(data_rows)


# ============================================================
# DISPATCHER
# ============================================================

_COUNTER_MAP = {
    ('ORDER',  'shopee'):           _count_shopee_order,
    ('ORDER',  'tiktok_tokopedia'): _count_tiktok_order,
    ('ORDER',  'lazada'):           _count_lazada_order,
    ('INCOME', 'shopee'):           _count_shopee_income,
    ('INCOME', 'tiktok_tokopedia'): _count_tiktok_income,
    ('INCOME', 'lazada'):           _count_lazada_income,
    ('REPORT', 'shopee'):           _count_shopee_report,
    ('REPORT', 'tiktok_tokopedia'): _count_tiktok_report,
    ('REPORT', 'lazada'):           _count_lazada_report,
}


def count_rows_in_file(file_path, fase, marketplace):
    """
    Menghitung jumlah baris data yang akan masuk ke tabel staging primer
    dari file spreadsheet yang diberikan.
    """
    key = (fase.upper(), marketplace.lower())
    counter_fn = _COUNTER_MAP.get(key)
    if counter_fn is None:
        logger.warning(f"Tidak ada counter untuk kombinasi fase={fase}, marketplace={marketplace}")
        return 0
    try:
        return counter_fn(file_path)
    except Exception as e:
        logger.error(f"Gagal menghitung baris dari file {file_path}: {e}")
        return 0


# ============================================================
# PENGECEKAN DUA LAPIS
# ============================================================

def check_file_status(filename, file_path, fase, marketplace, engine):
    """
    Melakukan pengecekan dua lapis untuk satu file:

    Lapis 1 — cek source_filename di tabel staging primer.
              Jika tidak ditemukan → status 'new', selesai.

    Lapis 2 — bandingkan jumlah baris di DB vs di file.
              Menghasilkan salah satu dari empat status:
              - 'new'          : belum pernah dimuat
              - 'fully_loaded' : seluruh baris sudah ada di DB
              - 'partial'      : sebagian baris sudah ada di DB
              - 'anomaly'      : baris di DB lebih banyak dari file

    Returns dict:
        {
            'status'       : str,
            'rows_in_db'   : int,
            'rows_in_file' : int,
            'table'        : str,
        }
    """
    key   = (fase.upper(), marketplace.lower())
    table = PRIMARY_TABLE.get(key)

    if table is None:
        return {'status': 'unknown', 'rows_in_db': 0, 'rows_in_file': 0, 'table': ''}

    # --- Lapis 1: cek keberadaan source_filename di DB ---
    try:
        with engine.connect() as conn:
            rows_in_db = conn.execute(
                text(f"SELECT COUNT(*) FROM {table} WHERE source_filename = :fn"),
                {"fn": filename}
            ).scalar()
    except Exception as e:
        logger.error(f"Gagal query DB saat check_file_status: {e}")
        return {'status': 'unknown', 'rows_in_db': 0, 'rows_in_file': 0, 'table': table}

    if rows_in_db == 0:
        return {'status': 'new', 'rows_in_db': 0, 'rows_in_file': 0, 'table': table}

    # --- Lapis 2: bandingkan jumlah baris ---
    rows_in_file = count_rows_in_file(file_path, fase, marketplace)

    if rows_in_db == rows_in_file:
        status = 'fully_loaded'
    elif rows_in_db < rows_in_file:
        status = 'partial'
    else:
        status = 'anomaly'

    return {
        'status'       : status,
        'rows_in_db'   : rows_in_db,
        'rows_in_file' : rows_in_file,
        'table'        : table,
    }


# ============================================================
# CEK DUPLIKASI PER ORDER ID
# ============================================================

def _read_order_ids_from_file(file_path, fase, marketplace):
    """Baca order ID dari file spreadsheet."""
    key = (fase.upper(), marketplace.lower())
    config = ORDER_ID_CONFIG.get(key)
    if not config:
        return set()

    excel_col = config[0]

    try:
        if fase.upper() == 'ORDER' and marketplace.lower() == 'tiktok_tokopedia':
            # TikTok ORDER: header di row 0, row 1 = deskripsi (skip), data mulai row 2
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            ws = wb.active
            rows = list(ws.values)
            wb.close()
            if len(rows) < 3:
                return set()
            header = [str(c).strip() if c else '' for c in rows[0]]
            if excel_col not in header:
                return set()
            idx = header.index(excel_col)
            return {str(r[idx]).strip() for r in rows[2:] if r[idx] is not None and str(r[idx]).strip() not in ('', 'nan')}

        elif fase.upper() == 'INCOME' and marketplace.lower() == 'shopee':
            # Shopee INCOME: baca semua sheet yang mengandung 'income', header di row 5
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            ids = set()
            for sheet in wb.sheetnames:
                if 'income' not in sheet.lower():
                    continue
                ws = wb[sheet]
                rows = list(ws.values)
                if len(rows) < 7:
                    continue
                header = [str(c).strip() if c else '' for c in rows[5]]
                if excel_col not in header:
                    continue
                idx = header.index(excel_col)
                for r in rows[6:]:
                    val = r[idx] if idx < len(r) else None
                    if val is not None and str(val).strip() not in ('', 'nan'):
                        ids.add(str(val).strip())
            wb.close()
            return ids

        elif fase.upper() == 'INCOME' and marketplace.lower() == 'tiktok_tokopedia':
            xl = pd.ExcelFile(file_path, engine='openpyxl')
            target = 'Order details' if 'Order details' in xl.sheet_names else 0
            df = pd.read_excel(file_path, sheet_name=target, dtype=str, engine='openpyxl')
            df.columns = df.columns.str.strip()
            if excel_col not in df.columns:
                return set()
            return set(df[excel_col].dropna().str.strip().replace('nan', '').replace('', pd.NA).dropna())

        elif fase.upper() == 'REPORT' and marketplace.lower() == 'shopee':
            # Shopee REPORT: sheet 'Transaction Report', header di row 17
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            if 'Transaction Report' not in wb.sheetnames:
                wb.close()
                return set()
            ws = wb['Transaction Report']
            rows = list(ws.values)
            wb.close()
            if len(rows) < 19:
                return set()
            header = [str(c).strip() if c else '' for c in rows[17]]
            if excel_col not in header:
                return set()
            idx = header.index(excel_col)
            return {str(r[idx]).strip() for r in rows[18:] if idx < len(r) and r[idx] is not None and str(r[idx]).strip() not in ('', 'nan')}

        else:
            # Default: read_excel biasa (Lazada ORDER/INCOME/REPORT, Shopee REPORT umum, TikTok REPORT)
            df = pd.read_excel(file_path, dtype=str, engine='openpyxl')
            df.columns = df.columns.str.strip()
            if excel_col not in df.columns:
                return set()
            return set(df[excel_col].dropna().str.strip().replace({'nan': None, '': None}).dropna())

    except Exception as e:
        logger.error(f"Gagal baca order ID dari {file_path}: {e}")
        return set()


def check_duplicate_order_ids(file_path, fase, marketplace, engine):
    """
    Cek order ID dari file terhadap staging database.

    Returns dict:
        {
            'total_in_file' : int,   jumlah order ID unik di file
            'already_in_db' : int,   jumlah yang sudah ada di staging
            'new'           : int,   jumlah yang belum ada
            'duplicate_ids' : list,  sample order ID yang duplikat (maks 10)
        }
    """
    key = (fase.upper(), marketplace.lower())
    config = ORDER_ID_CONFIG.get(key)
    if not config:
        return {'total_in_file': 0, 'already_in_db': 0, 'new': 0, 'duplicate_ids': []}

    _, staging_col, staging_table = config

    ids_in_file = _read_order_ids_from_file(file_path, fase, marketplace)
    if not ids_in_file:
        return {'total_in_file': 0, 'already_in_db': 0, 'new': 0, 'duplicate_ids': []}

    try:
        with engine.connect() as conn:
            result = conn.execute(text(f"""
                SELECT {staging_col}
                FROM staging.{staging_table}
                WHERE {staging_col} = ANY(:ids)
            """), {'ids': list(ids_in_file)})
            ids_in_db = {r[0] for r in result}

        duplicates  = ids_in_file & ids_in_db
        new_ids     = ids_in_file - ids_in_db

        return {
            'total_in_file' : len(ids_in_file),
            'already_in_db' : len(duplicates),
            'new'           : len(new_ids),
            'duplicate_ids' : sorted(list(duplicates))[:10],
        }
    except Exception as e:
        logger.error(f"Gagal cek duplikat order ID: {e}")
        return {'total_in_file': len(ids_in_file), 'already_in_db': 0, 'new': len(ids_in_file), 'duplicate_ids': []}
